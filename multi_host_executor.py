"""
Multi-Host SSH Command Executor
Execute same commands on multiple remote systems simultaneously
Enhanced Modern UI Design

Enhancements:
- SSH Key Authentication support
- Stop/Cancel execution
- Re-run failed hosts
- Import hosts from file (TXT/CSV)
- Configurable max concurrent connections
- Copy results to clipboard
- Session config save/load
- Host groups/profiles support
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import threading
import queue
from concurrent.futures import ThreadPoolExecutor, as_completed
import paramiko
import socket
from datetime import datetime
import json
import os
import csv
import shlex
import hashlib
import secrets
import sys

# Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ============== TOOLTIP ==============
class Tooltip:
    """Lightweight hover tooltip for any widget."""

    def __init__(self, widget, text, delay=450):
        self.widget = widget
        self.text = text
        self.delay = delay
        self._tip = None
        self._after_id = None
        widget.bind('<Enter>', self._schedule, add='+')
        widget.bind('<Leave>', self._hide, add='+')
        widget.bind('<ButtonPress>', self._hide, add='+')

    def _schedule(self, _event=None):
        self._cancel()
        self._after_id = self.widget.after(self.delay, self._show)

    def _cancel(self):
        if self._after_id:
            try:
                self.widget.after_cancel(self._after_id)
            except Exception:
                pass
            self._after_id = None

    def _show(self):
        if self._tip or not self.text:
            return
        x = self.widget.winfo_rootx() + 16
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 6
        tip = tk.Toplevel(self.widget)
        tip.wm_overrideredirect(True)
        tip.wm_geometry(f"+{x}+{y}")
        tip.configure(bg="#0f172a")
        frame = tk.Frame(tip, bg="#0f172a", highlightbackground="#7c3aed",
                         highlightthickness=1)
        frame.pack()
        tk.Label(frame, text=self.text, font=('Segoe UI', 9),
                 bg="#0f172a", fg="#e2e8f0",
                 padx=10, pady=6, justify='left', wraplength=320).pack()
        self._tip = tip

    def _hide(self, _event=None):
        self._cancel()
        if self._tip:
            try:
                self._tip.destroy()
            except Exception:
                pass
            self._tip = None


# ============== ADMIN AUTH MANAGER ==============
class AdminAuthManager:
    """
    Protects access to the application with an admin password.
    Stores PBKDF2-HMAC-SHA256 hash + salt in JSON. Plain password is never persisted.
    """

    HASH_ITERATIONS = 200_000
    SALT_BYTES = 32

    def __init__(self):
        self.auth_file = os.path.join(os.path.expanduser("~"), ".ssh_executor_admin.json")

    def is_configured(self):
        return os.path.exists(self.auth_file)

    def _hash(self, password, salt):
        return hashlib.pbkdf2_hmac(
            'sha256', password.encode('utf-8'), salt, self.HASH_ITERATIONS
        ).hex()

    def set_password(self, password):
        salt = secrets.token_bytes(self.SALT_BYTES)
        data = {
            'salt': salt.hex(),
            'hash': self._hash(password, salt),
            'iterations': self.HASH_ITERATIONS,
            'algo': 'pbkdf2_sha256',
        }
        try:
            with open(self.auth_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
            try:
                os.chmod(self.auth_file, 0o600)
            except Exception:
                pass
            return True
        except Exception:
            return False

    def verify(self, password):
        try:
            with open(self.auth_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            salt = bytes.fromhex(data['salt'])
            expected = data['hash']
            iterations = data.get('iterations', self.HASH_ITERATIONS)
            actual = hashlib.pbkdf2_hmac(
                'sha256', password.encode('utf-8'), salt, iterations
            ).hex()
            return secrets.compare_digest(actual, expected)
        except Exception:
            return False

    def reset(self):
        try:
            if os.path.exists(self.auth_file):
                os.remove(self.auth_file)
            return True
        except Exception:
            return False


# ============== ADMIN LOGIN DIALOG ==============
class AdminLoginDialog:
    """Modal dialog that gates app access behind an admin password."""

    MAX_ATTEMPTS = 3

    def __init__(self, parent, auth_manager):
        self.auth = auth_manager
        self.authenticated = False
        self.attempts_left = self.MAX_ATTEMPTS
        self.first_run = not self.auth.is_configured()

        self.win = tk.Toplevel(parent)
        self.win.title("Admin Authentication" if not self.first_run else "Admin Setup")
        self.win.configure(bg="#1a1a2e")
        self.win.resizable(False, False)
        self.win.transient(parent)
        self.win.grab_set()
        self.win.protocol("WM_DELETE_WINDOW", self._on_cancel)

        self._build_ui()
        self._center(parent, 440, 320 if self.first_run else 280)

    def _center(self, parent, w, h):
        parent.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - w) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        if x < 0 or y < 0:
            sw = parent.winfo_screenwidth()
            sh = parent.winfo_screenheight()
            x = (sw - w) // 2
            y = (sh - h) // 2
        self.win.geometry(f"{w}x{h}+{x}+{y}")

    def _build_ui(self):
        # Header bar
        header = tk.Frame(self.win, bg="#7c3aed", height=6)
        header.pack(fill=tk.X)

        body = tk.Frame(self.win, bg="#1a1a2e", padx=28, pady=22)
        body.pack(fill=tk.BOTH, expand=True)

        # Lock icon + title
        title_row = tk.Frame(body, bg="#1a1a2e")
        title_row.pack(fill=tk.X)
        tk.Label(title_row, text="🔒", font=('Segoe UI', 22),
                 bg="#1a1a2e", fg="#ffd166").pack(side=tk.LEFT, padx=(0, 10))
        title_text = "Set Admin Password" if self.first_run else "Admin Login Required"
        tk.Label(title_row, text=title_text, font=('Segoe UI', 14, 'bold'),
                 bg="#1a1a2e", fg="#ffffff").pack(side=tk.LEFT, anchor='w')

        subtitle = (
            "Create a password to protect this application.\nIt will be required every time you launch."
            if self.first_run
            else "Only administrators with the password can open this app."
        )
        tk.Label(body, text=subtitle, font=('Segoe UI', 9),
                 bg="#1a1a2e", fg="#94a3b8", justify='left').pack(anchor='w', pady=(8, 14))

        # Password entry
        tk.Label(body, text="Password", font=('Segoe UI', 9, 'bold'),
                 bg="#1a1a2e", fg="#cbd5e1").pack(anchor='w')
        self.pw_entry = tk.Entry(body, show='•', font=('Consolas', 12),
                                 bg="#0f172a", fg="#ffffff", relief=tk.FLAT,
                                 insertbackground="#ffffff")
        self.pw_entry.pack(fill=tk.X, ipady=7, pady=(4, 8))

        if self.first_run:
            tk.Label(body, text="Confirm Password", font=('Segoe UI', 9, 'bold'),
                     bg="#1a1a2e", fg="#cbd5e1").pack(anchor='w')
            self.pw_confirm = tk.Entry(body, show='•', font=('Consolas', 12),
                                       bg="#0f172a", fg="#ffffff", relief=tk.FLAT,
                                       insertbackground="#ffffff")
            self.pw_confirm.pack(fill=tk.X, ipady=7, pady=(4, 8))
        else:
            self.pw_confirm = None

        # Status label
        self.status_label = tk.Label(body, text="", font=('Segoe UI', 9),
                                     bg="#1a1a2e", fg="#ef4444")
        self.status_label.pack(anchor='w', pady=(4, 6))

        # Buttons
        btn_row = tk.Frame(body, bg="#1a1a2e")
        btn_row.pack(fill=tk.X, pady=(8, 0))

        ok_text = "Create Password" if self.first_run else "Unlock"
        tk.Button(btn_row, text=ok_text, font=('Segoe UI', 10, 'bold'),
                  bg="#10b981", fg="#ffffff", activebackground="#34d399",
                  relief=tk.FLAT, cursor='hand2', width=16,
                  command=self._on_submit).pack(side=tk.RIGHT, padx=(6, 0), ipady=5)

        tk.Button(btn_row, text="Cancel", font=('Segoe UI', 10, 'bold'),
                  bg="#475569", fg="#ffffff", activebackground="#64748b",
                  relief=tk.FLAT, cursor='hand2', width=10,
                  command=self._on_cancel).pack(side=tk.RIGHT, ipady=5)

        self.pw_entry.focus_set()
        self.win.bind('<Return>', lambda e: self._on_submit())
        self.win.bind('<Escape>', lambda e: self._on_cancel())

    def _set_status(self, text, color="#ef4444"):
        self.status_label.config(text=text, fg=color)

    def _on_submit(self):
        password = self.pw_entry.get()

        if self.first_run:
            confirm = self.pw_confirm.get()
            if len(password) < 6:
                self._set_status("Password must be at least 6 characters.")
                return
            if password != confirm:
                self._set_status("Passwords do not match.")
                self.pw_confirm.delete(0, tk.END)
                self.pw_confirm.focus_set()
                return
            if self.auth.set_password(password):
                self.authenticated = True
                self.win.destroy()
            else:
                self._set_status("Failed to save password file.")
            return

        if self.auth.verify(password):
            self.authenticated = True
            self.win.destroy()
        else:
            self.attempts_left -= 1
            self.pw_entry.delete(0, tk.END)
            if self.attempts_left <= 0:
                self._set_status("Too many failed attempts. Exiting.")
                self.win.after(1200, self._on_cancel)
            else:
                self._set_status(
                    f"Incorrect password. {self.attempts_left} attempt(s) left."
                )

    def _on_cancel(self):
        self.authenticated = False
        try:
            self.win.destroy()
        except Exception:
            pass


# ============== COMMAND PRESETS MANAGER ==============
class CommandPresetManager:
    """Manages saving and loading command presets"""

    def __init__(self):
        self.presets_file = os.path.join(os.path.expanduser("~"), ".ssh_executor_presets.json")
        self.presets = self.load_presets()

    def load_presets(self):
        try:
            if os.path.exists(self.presets_file):
                with open(self.presets_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception:
            pass
        return {}

    def save_presets(self):
        try:
            with open(self.presets_file, 'w', encoding='utf-8') as f:
                json.dump(self.presets, f, indent=2)
            return True
        except Exception:
            return False

    def add_preset(self, name, commands):
        self.presets[name] = commands
        return self.save_presets()

    def delete_preset(self, name):
        if name in self.presets:
            del self.presets[name]
            return self.save_presets()
        return False

    def get_preset(self, name):
        return self.presets.get(name, [])

    def get_preset_names(self):
        return list(self.presets.keys())


# ============== HOST GROUPS MANAGER ==============
class HostGroupManager:
    """Manages saving and loading host groups"""

    def __init__(self):
        self.groups_file = os.path.join(os.path.expanduser("~"), ".ssh_executor_host_groups.json")
        self.groups = self.load_groups()

    def load_groups(self):
        try:
            if os.path.exists(self.groups_file):
                with open(self.groups_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception:
            pass
        return {}

    def save_groups(self):
        try:
            with open(self.groups_file, 'w', encoding='utf-8') as f:
                json.dump(self.groups, f, indent=2)
            return True
        except Exception:
            return False

    def add_group(self, name, hosts):
        self.groups[name] = hosts
        return self.save_groups()

    def delete_group(self, name):
        if name in self.groups:
            del self.groups[name]
            return self.save_groups()
        return False

    def get_group(self, name):
        return self.groups.get(name, [])

    def get_group_names(self):
        return list(self.groups.keys())


# ============== SESSION CONFIG MANAGER ==============
class SessionConfigManager:
    """Manages saving and loading session configurations"""

    def __init__(self):
        self.config_file = os.path.join(os.path.expanduser("~"), ".ssh_executor_session.json")

    def save_session(self, config):
        """Save session config (excluding password)"""
        try:
            # Don't save password
            safe_config = {k: v for k, v in config.items() if k != 'password'}
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(safe_config, f, indent=2)
            return True
        except Exception:
            return False

    def load_session(self):
        """Load session config"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception:
            pass
        return {}


# ============== ULTRA COLORFUL MODERN THEME ==============
class Theme:
    # Main backgrounds - Softer, warmer dark tones
    BG_DARK = "#0f1318"
    BG_MAIN = "#161b22"
    BG_CARD = "#1e252e"
    BG_INPUT = "#262f3d"
    BG_OUTPUT = "#121820"
    BG_HOVER = "#2d3848"

    # Gradient colors for header - Softer gradient
    HEADER_START = "#7c8ce8"
    HEADER_MID = "#8b6bab"
    HEADER_END = "#d9a0e0"
    HEADER_ACCENT = "#fff5eb"

    # Animated gradient colors - Softer palette
    GRADIENT_1 = "#7c8ce8"
    GRADIENT_2 = "#8b6bab"
    GRADIENT_3 = "#d9a0e0"
    GRADIENT_4 = "#e8828f"
    GRADIENT_5 = "#6fb8f0"

    # Accent colors - Softer Blue/Purple
    PRIMARY = "#7477f0"
    PRIMARY_HOVER = "#9094f5"
    PRIMARY_DARK = "#5c5ee0"
    PRIMARY_GLOW = "#d4d8fc"

    SECONDARY = "#b279e8"
    SECONDARY_HOVER = "#c99df0"
    SECONDARY_DARK = "#9654d8"
    SECONDARY_GLOW = "#eddff8"

    # Status colors - Softer, less neon
    SUCCESS = "#2ecc87"
    SUCCESS_LIGHT = "#5dd9a5"
    SUCCESS_BG = "#0d2a20"
    SUCCESS_BORDER = "#25a770"
    SUCCESS_GLOW = "#8be5c0"
    SUCCESS_NEON = "#4aeaa0"

    ERROR = "#e86b6b"
    ERROR_LIGHT = "#f09090"
    ERROR_BG = "#2a1418"
    ERROR_BORDER = "#d45050"
    ERROR_GLOW = "#f5b5b5"
    ERROR_NEON = "#f07080"

    WARNING = "#e8a830"
    WARNING_LIGHT = "#f0c050"
    WARNING_BG = "#2a2210"
    WARNING_BORDER = "#d09020"
    WARNING_NEON = "#f0c040"

    INFO = "#40c4d8"
    INFO_LIGHT = "#60d8e8"
    INFO_BG = "#102428"
    INFO_BORDER = "#30a8c0"
    INFO_NEON = "#50d8f0"

    # Accent Colors - Softer spectrum
    CYAN = "#40c4d8"
    CYAN_LIGHT = "#60d8e8"
    CYAN_NEON = "#70e8f8"

    PINK = "#e070a8"
    PINK_LIGHT = "#e890c0"
    PINK_NEON = "#f0a0c8"

    ORANGE = "#e88848"
    ORANGE_LIGHT = "#f0a060"
    ORANGE_NEON = "#f0a068"

    LIME = "#90c830"
    LIME_LIGHT = "#a8d850"
    LIME_NEON = "#80e870"

    PURPLE = "#9878e8"
    PURPLE_LIGHT = "#b098f0"
    PURPLE_NEON = "#b090f0"

    GOLD = "#e0b820"
    GOLD_LIGHT = "#f0c830"
    GOLD_NEON = "#f0d050"

    CORAL = "#e89098"
    CORAL_LIGHT = "#f0b0b8"

    TEAL = "#38c0b0"
    TEAL_LIGHT = "#58d0c0"

    INDIGO = "#7477f0"
    INDIGO_LIGHT = "#9094f5"

    # Section-specific colors - Softer
    SECTION_HOSTS = "#40c4d8"      # Softer cyan for hosts
    SECTION_CREDS = "#e8a830"      # Softer amber for credentials
    SECTION_CMDS = "#9878e8"       # Softer purple for commands
    SECTION_RESULTS = "#2ecc87"    # Softer green for results

    # Text colors - Slightly softer white
    TEXT_WHITE = "#f8fafc"
    TEXT_LIGHT = "#e8eef4"
    TEXT_MEDIUM = "#a0aec0"
    TEXT_MUTED = "#718096"
    TEXT_DARK = "#2d3748"

    # Borders - Softer contrast
    BORDER = "#3d4a5c"
    BORDER_LIGHT = "#4a5a6e"
    BORDER_FOCUS = "#7477f0"
    BORDER_GLOW = "#b0b8e0"

    # Button specific colors - Softer
    BTN_EXECUTE = "#2ecc87"
    BTN_EXECUTE_HOVER = "#4dd8a0"
    BTN_STOP = "#e86b6b"
    BTN_STOP_HOVER = "#f09090"
    BTN_RETRY = "#e8a830"
    BTN_RETRY_HOVER = "#f0c050"
    BTN_CLEAR = "#718096"
    BTN_CLEAR_HOVER = "#a0aec0"

    # Animation timing (in milliseconds)
    ANIM_FAST = 150
    ANIM_NORMAL = 250
    ANIM_SLOW = 400

    # Spacing constants
    PAD_XS = 4
    PAD_SM = 8
    PAD_MD = 12
    PAD_LG = 16
    PAD_XL = 24
    PAD_XXL = 32


# ============== STUNNING HEADER WITH GLOW EFFECT ==============
class StunningHeader(tk.Frame):
    """Beautiful header with glowing animated border and modern design"""

    def __init__(self, parent):
        super().__init__(parent, bg=Theme.BG_DARK)
        self.pack(fill=tk.X)

        # Softer glow colors for pleasant feel
        self.glow_colors = [
            Theme.GRADIENT_1, Theme.GRADIENT_2, Theme.GRADIENT_3,
            Theme.GRADIENT_4, Theme.GRADIENT_5, Theme.TEAL
        ]
        self.glow_index = 0
        self.transition_step = 0
        self.transition_steps = 10  # Smoother transition

        # Glowing top border - slightly thicker for better visibility
        self.glow_border = tk.Frame(self, height=3, bg=self.glow_colors[0])
        self.glow_border.pack(fill=tk.X)

        # Main header container with softer background
        header_bg = tk.Frame(self, bg=Theme.BG_CARD, height=85)
        header_bg.pack(fill=tk.X)
        header_bg.pack_propagate(False)

        # Inner content with better padding
        content = tk.Frame(header_bg, bg=Theme.BG_CARD)
        content.pack(expand=True, fill=tk.BOTH, padx=Theme.PAD_XXL)

        # Left side - Logo and title
        left_side = tk.Frame(content, bg=Theme.BG_CARD)
        left_side.pack(side=tk.LEFT, fill=tk.Y, pady=Theme.PAD_MD)

        # Animated logo box with rounded feel
        self.logo_frame = tk.Frame(left_side, bg=Theme.PRIMARY, width=50, height=50)
        self.logo_frame.pack(side=tk.LEFT)
        self.logo_frame.pack_propagate(False)

        self.logo_label = tk.Label(self.logo_frame, text="SSH", font=('Consolas', 13, 'bold'),
                                   bg=Theme.PRIMARY, fg=Theme.TEXT_WHITE)
        self.logo_label.place(relx=0.5, rely=0.5, anchor='center')

        # Title section with better spacing
        title_frame = tk.Frame(left_side, bg=Theme.BG_CARD)
        title_frame.pack(side=tk.LEFT, padx=Theme.PAD_LG)

        # Main title with softer gradient-like effect
        title_row = tk.Frame(title_frame, bg=Theme.BG_CARD)
        title_row.pack(anchor=tk.W)

        tk.Label(title_row, text="MultiHost ", font=('Segoe UI', 20, 'bold'),
                bg=Theme.BG_CARD, fg=Theme.PRIMARY).pack(side=tk.LEFT)
        tk.Label(title_row, text="SSH", font=('Segoe UI', 20, 'bold'),
                bg=Theme.BG_CARD, fg=Theme.PINK).pack(side=tk.LEFT)

        # Subtitle with better contrast
        tk.Label(title_frame, text="Run commands on many servers at once — simple, fast, secure",
                font=('Segoe UI', 10), bg=Theme.BG_CARD, fg=Theme.TEXT_MEDIUM).pack(anchor=tk.W, pady=(4, 0))

        # Right side - Status indicators
        right_side = tk.Frame(content, bg=Theme.BG_CARD)
        right_side.pack(side=tk.RIGHT, fill=tk.Y, pady=Theme.PAD_LG)

        # Version badge with softer color
        version_frame = tk.Frame(right_side, bg=Theme.SUCCESS)
        version_frame.pack(side=tk.RIGHT, padx=Theme.PAD_SM)
        tk.Label(version_frame, text=" v2.0 ", font=('Segoe UI', 9, 'bold'),
                bg=Theme.SUCCESS, fg=Theme.TEXT_WHITE).pack(padx=10, pady=5)

        # Status indicator
        self.status_dot = tk.Frame(right_side, bg=Theme.SUCCESS, width=10, height=10)
        self.status_dot.pack(side=tk.RIGHT, padx=Theme.PAD_MD)

        tk.Label(right_side, text="Ready", font=('Segoe UI', 10),
                bg=Theme.BG_CARD, fg=Theme.SUCCESS).pack(side=tk.RIGHT, padx=Theme.PAD_SM)

        # Bottom accent line - softer
        bottom_accent = tk.Frame(self, height=1, bg=Theme.BORDER)
        bottom_accent.pack(fill=tk.X)

        # Start smooth animations
        self.animate_glow_smooth()

    def _lerp_color(self, color1, color2, t):
        """Linear interpolate between two hex colors"""
        r1, g1, b1 = int(color1[1:3], 16), int(color1[3:5], 16), int(color1[5:7], 16)
        r2, g2, b2 = int(color2[1:3], 16), int(color2[3:5], 16), int(color2[5:7], 16)
        r = int(r1 + (r2 - r1) * t)
        g = int(g1 + (g2 - g1) * t)
        b = int(b1 + (b2 - b1) * t)
        return f"#{r:02x}{g:02x}{b:02x}"

    def animate_glow_smooth(self):
        """Animate with smooth color transitions"""
        current_color = self.glow_colors[self.glow_index]
        next_index = (self.glow_index + 1) % len(self.glow_colors)
        next_color = self.glow_colors[next_index]

        t = self.transition_step / self.transition_steps
        blended = self._lerp_color(current_color, next_color, t)

        self.glow_border.config(bg=blended)
        self.logo_frame.config(bg=blended)
        self.logo_label.config(bg=blended)

        self.transition_step += 1
        if self.transition_step > self.transition_steps:
            self.transition_step = 0
            self.glow_index = next_index

        self.after(80, self.animate_glow_smooth)  # Smoother ~12fps transition


# ============== COLORFUL SECTION HEADER ==============
class ColorfulSectionHeader(tk.Frame):
    """Section header with icon and accent color"""

    def __init__(self, parent, title, subtitle="", icon="", accent_color=None):
        super().__init__(parent, bg=Theme.BG_MAIN)
        self.pack(fill=tk.X, pady=(Theme.PAD_MD, Theme.PAD_SM))

        accent = accent_color or Theme.PRIMARY

        # Container with subtle left border accent
        container = tk.Frame(self, bg=Theme.BG_MAIN)
        container.pack(fill=tk.X)

        # Colored accent bar on the left - slightly thinner for elegance
        accent_bar = tk.Frame(container, bg=accent, width=3)
        accent_bar.pack(side=tk.LEFT, fill=tk.Y, padx=(0, Theme.PAD_LG))

        # Icon (if provided)
        if icon:
            icon_bg = tk.Frame(container, bg=accent, width=34, height=34)
            icon_bg.pack(side=tk.LEFT, padx=(0, Theme.PAD_MD))
            icon_bg.pack_propagate(False)
            tk.Label(icon_bg, text=icon, font=('Segoe UI', 12, 'bold'),
                    bg=accent, fg=Theme.TEXT_WHITE).place(relx=0.5, rely=0.5, anchor='center')

        # Text container
        text_frame = tk.Frame(container, bg=Theme.BG_MAIN)
        text_frame.pack(side=tk.LEFT, fill=tk.X)

        # Title with accent color
        tk.Label(text_frame, text=title, font=('Segoe UI', 12, 'bold'),
                bg=Theme.BG_MAIN, fg=accent).pack(anchor=tk.W)

        # Subtitle with slightly more spacing
        if subtitle:
            tk.Label(text_frame, text=subtitle, font=('Segoe UI', 9),
                    bg=Theme.BG_MAIN, fg=Theme.TEXT_MUTED).pack(anchor=tk.W, pady=(2, 0))


# ============== MODERN BUTTON CLASS ==============
class ModernButton(tk.Frame):
    """Modern flat button with smooth hover effects"""

    def __init__(self, parent, text, command, bg_color, hover_color,
                 fg_color="#ffffff", width=100, height=40, font_size=10, icon=""):
        super().__init__(parent, bg=parent.cget('bg') if hasattr(parent, 'cget') else Theme.BG_MAIN)

        self.command = command
        self.bg_color = bg_color
        self.hover_color = hover_color
        self.fg_color = fg_color
        self.enabled = True
        self._animation_id = None
        self._current_color = bg_color

        # Button frame with colored background and better padding
        self.btn_frame = tk.Frame(self, bg=bg_color, cursor='hand2')
        self.btn_frame.pack(padx=3, pady=3)

        # Button label with improved padding
        display_text = f"{icon} {text}" if icon else text
        self.label = tk.Label(self.btn_frame, text=display_text,
                             font=('Segoe UI', font_size, 'bold'),
                             bg=bg_color, fg=fg_color,
                             width=width//10, height=1,
                             cursor='hand2')
        self.label.pack(padx=Theme.PAD_LG, pady=Theme.PAD_SM)

        # Bind events
        for widget in [self.btn_frame, self.label]:
            widget.bind('<Enter>', self.on_enter)
            widget.bind('<Leave>', self.on_leave)
            widget.bind('<Button-1>', self.on_click)

    def _lerp_color(self, color1, color2, t):
        """Linear interpolate between two hex colors"""
        try:
            r1, g1, b1 = int(color1[1:3], 16), int(color1[3:5], 16), int(color1[5:7], 16)
            r2, g2, b2 = int(color2[1:3], 16), int(color2[3:5], 16), int(color2[5:7], 16)
            r = int(r1 + (r2 - r1) * t)
            g = int(g1 + (g2 - g1) * t)
            b = int(b1 + (b2 - b1) * t)
            return f"#{r:02x}{g:02x}{b:02x}"
        except Exception:
            return color2

    def _animate_to_color(self, target_color, steps=5, current_step=0):
        """Smoothly animate to target color"""
        if self._animation_id:
            self.after_cancel(self._animation_id)

        if current_step <= steps:
            t = current_step / steps
            color = self._lerp_color(self._current_color, target_color, t)
            self.btn_frame.config(bg=color)
            self.label.config(bg=color)
            if current_step == steps:
                self._current_color = target_color
            self._animation_id = self.after(20, lambda: self._animate_to_color(target_color, steps, current_step + 1))

    def on_enter(self, event):
        if self.enabled:
            self._animate_to_color(self.hover_color)

    def on_leave(self, event):
        if self.enabled:
            self._animate_to_color(self.bg_color)

    def on_click(self, event):
        if self.enabled and self.command:
            self.command()

    def set_enabled(self, enabled):
        self.enabled = enabled
        if enabled:
            self._current_color = self.bg_color
            self.btn_frame.config(bg=self.bg_color)
            self.label.config(bg=self.bg_color, fg=self.fg_color)
            self.btn_frame.config(cursor='hand2')
            self.label.config(cursor='hand2')
        else:
            self._current_color = Theme.TEXT_MUTED
            self.btn_frame.config(bg=Theme.TEXT_MUTED)
            self.label.config(bg=Theme.TEXT_MUTED, fg=Theme.TEXT_DARK)
            self.btn_frame.config(cursor='')
            self.label.config(cursor='')


# ============== STYLED BUTTON CLASS (Legacy) ==============
class StyledButton(tk.Canvas):
    """Modern styled button with hover effects and glow"""

    def __init__(self, parent, text, command, bg_color, hover_color,
                 fg_color="#ffffff", width=100, height=36, font_size=10, icon="",
                 glow_color=None, style="default"):
        # Get parent background color
        try:
            parent_bg = parent.cget('bg')
        except Exception:
            parent_bg = Theme.BG_MAIN
        super().__init__(parent, width=width+4, height=height+4,
                        bg=parent_bg, highlightthickness=0)

        self.command = command
        self.bg_color = bg_color
        self.hover_color = hover_color
        self.fg_color = fg_color
        self.text = text
        self.icon = icon
        self.btn_width = width
        self.btn_height = height
        self.font_size = font_size
        self.enabled = True
        self.glow_color = glow_color or hover_color
        self.style = style
        self.is_hovered = False

        self.draw_button(bg_color, False)

        self.bind('<Enter>', self.on_enter)
        self.bind('<Leave>', self.on_leave)
        self.bind('<Button-1>', self.on_click)
        self.bind('<ButtonRelease-1>', self.on_release)

    def draw_button(self, color, with_glow=False):
        self.delete("all")

        offset = 2
        r = 10  # radius for more rounded corners
        w = self.btn_width
        h = self.btn_height

        # Draw glow/shadow effect when hovered
        if with_glow and self.enabled:
            glow_color = self.glow_color
            # Outer glow
            self.create_oval(offset-2, offset-2, r*2+offset-2, r*2+offset-2, fill=glow_color, outline="")
            self.create_oval(w-r*2+offset+2, offset-2, w+offset+2, r*2+offset-2, fill=glow_color, outline="")
            self.create_oval(offset-2, h-r*2+offset+2, r*2+offset-2, h+offset+2, fill=glow_color, outline="")
            self.create_oval(w-r*2+offset+2, h-r*2+offset+2, w+offset+2, h+offset+2, fill=glow_color, outline="")
            self.create_rectangle(r+offset-2, offset-2, w-r+offset+2, h+offset+2, fill=glow_color, outline="")
            self.create_rectangle(offset-2, r+offset-2, w+offset+2, h-r+offset+2, fill=glow_color, outline="")

        # Draw main button shape
        self.create_oval(offset, offset, r*2+offset, r*2+offset, fill=color, outline="")
        self.create_oval(w-r*2+offset, offset, w+offset, r*2+offset, fill=color, outline="")
        self.create_oval(offset, h-r*2+offset, r*2+offset, h+offset, fill=color, outline="")
        self.create_oval(w-r*2+offset, h-r*2+offset, w+offset, h+offset, fill=color, outline="")
        self.create_rectangle(r+offset, offset, w-r+offset, h+offset, fill=color, outline="")
        self.create_rectangle(offset, r+offset, w+offset, h-r+offset, fill=color, outline="")

        # Draw subtle top highlight for 3D effect
        if self.enabled:
            highlight = self._lighten_color(color, 30)
            self.create_line(r+offset, offset+1, w-r+offset, offset+1, fill=highlight, width=1)

        # Draw text with shadow for depth
        display_text = f"{self.icon} {self.text}" if self.icon else self.text

        # Text shadow
        if self.enabled:
            shadow_color = self._darken_color(color, 50)
            self.create_text((w+offset*2)//2+1, (h+offset*2)//2+1, text=display_text,
                            fill=shadow_color, font=('Segoe UI', self.font_size, 'bold'))

        # Main text
        self.create_text((w+offset*2)//2, (h+offset*2)//2, text=display_text,
                        fill=self.fg_color, font=('Segoe UI', self.font_size, 'bold'))

    def _lighten_color(self, hex_color, amount):
        """Lighten a hex color"""
        try:
            hex_color = hex_color.lstrip('#')
            r = min(255, int(hex_color[0:2], 16) + amount)
            g = min(255, int(hex_color[2:4], 16) + amount)
            b = min(255, int(hex_color[4:6], 16) + amount)
            return f'#{r:02x}{g:02x}{b:02x}'
        except Exception:
            return hex_color

    def _darken_color(self, hex_color, amount):
        """Darken a hex color"""
        try:
            hex_color = hex_color.lstrip('#')
            r = max(0, int(hex_color[0:2], 16) - amount)
            g = max(0, int(hex_color[2:4], 16) - amount)
            b = max(0, int(hex_color[4:6], 16) - amount)
            return f'#{r:02x}{g:02x}{b:02x}'
        except Exception:
            return hex_color

    def on_enter(self, event):
        if self.enabled:
            self.is_hovered = True
            self.draw_button(self.hover_color, True)
            self.config(cursor='hand2')

    def on_leave(self, event):
        if self.enabled:
            self.is_hovered = False
            self.draw_button(self.bg_color, False)

    def on_click(self, event):
        if self.enabled:
            # Draw pressed state
            self.draw_button(self._darken_color(self.bg_color, 20), False)
            # Call command on click
            if self.command:
                self.after(50, self.command)  # Small delay for visual feedback

    def on_release(self, event):
        if self.enabled:
            self.draw_button(self.hover_color if self.is_hovered else self.bg_color, self.is_hovered)

    def set_enabled(self, enabled):
        self.enabled = enabled
        if enabled:
            self.draw_button(self.bg_color, False)
        else:
            self.draw_button(Theme.TEXT_MUTED, False)


# ============== SSH EXECUTOR ==============
class SSHExecutor:
    def __init__(self, hostname, username, password=None, port=22, timeout=30,
                 key_file=None, key_passphrase=None):
        self.hostname = hostname
        self.username = username
        self.password = password
        self.port = port
        self.timeout = timeout
        self.key_file = key_file
        self.key_passphrase = key_passphrase
        self.cancelled = False
        self._client = None

    def cancel(self):
        self.cancelled = True
        # Forcibly close the active SSH client to interrupt blocking I/O
        client = self._client
        if client is not None:
            try:
                client.close()
            except Exception:
                pass

    def execute_commands(self, commands):
        results = {
            'hostname': self.hostname,
            'success': False,
            'status': 'FAILED',
            'output': '',
            'error': '',
            'commands_output': [],
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        if self.cancelled:
            results['error'] = "Execution cancelled"
            results['status'] = 'CANCELLED'
            return results

        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        self._client = client

        try:
            # Prepare connection arguments
            connect_args = {
                'hostname': self.hostname,
                'port': self.port,
                'username': self.username,
                'timeout': self.timeout,
                'look_for_keys': False,
                'allow_agent': False,
                'banner_timeout': self.timeout
            }

            # Use SSH key if provided
            if self.key_file and os.path.exists(self.key_file):
                try:
                    # Try different key types
                    key = None
                    for key_class in [paramiko.RSAKey, paramiko.Ed25519Key, paramiko.ECDSAKey, paramiko.DSSKey]:
                        try:
                            key = key_class.from_private_key_file(
                                self.key_file,
                                password=self.key_passphrase
                            )
                            break
                        except Exception:
                            continue

                    if key:
                        connect_args['pkey'] = key
                    else:
                        # Fall back to password
                        if self.password:
                            connect_args['password'] = self.password
                except Exception as e:
                    results['error'] = f"Key file error: {str(e)}"
                    results['status'] = 'KEY ERROR'
                    return results
            elif self.password:
                connect_args['password'] = self.password
            else:
                results['error'] = "No authentication method provided"
                results['status'] = 'AUTH ERROR'
                return results

            client.connect(**connect_args)

            if self.cancelled:
                results['error'] = "Execution cancelled"
                results['status'] = 'CANCELLED'
                return results

            commands_output = []
            valid_commands = [cmd.strip() for cmd in commands if cmd.strip()]

            if valid_commands:
                marker = "__OUTPUT_MARKER_98765__"
                cmd_parts = []
                for cmd in valid_commands:
                    # Handle sudo commands - use -S flag to read password from stdin
                    if cmd.strip().startswith('sudo ') and self.password:
                        sudo_cmd = cmd.strip()[5:]  # Remove 'sudo ' prefix
                        # shlex.quote safely escapes any character (backticks, $, !, etc.)
                        quoted_pass = shlex.quote(self.password)
                        cmd = f"printf '%s\\n' {quoted_pass} | sudo -S -p '' {sudo_cmd}"
                    cmd_parts.append(cmd)
                    cmd_parts.append(f'echo "{marker}"')

                combined_cmd = '; '.join(cmd_parts)

                stdin, stdout, stderr = client.exec_command(
                    combined_cmd,
                    timeout=self.timeout * len(valid_commands),
                    get_pty=True  # Use pseudo-terminal for interactive commands
                )

                full_output = stdout.read().decode('utf-8', errors='replace')
                full_error = stderr.read().decode('utf-8', errors='replace')

                output_parts = full_output.split(marker)

                for i, cmd in enumerate(valid_commands):
                    if i < len(output_parts):
                        output = output_parts[i].strip()
                    else:
                        output = ''

                    commands_output.append({
                        'command': cmd,
                        'output': output if output else '(command executed)',
                        'error': '',
                        'exit_code': 0
                    })

                if full_error.strip() and commands_output:
                    commands_output[-1]['error'] = full_error.strip()

            results['commands_output'] = commands_output
            results['success'] = True
            results['status'] = 'DONE'

        except paramiko.AuthenticationException:
            results['error'] = "Authentication failed - Wrong username or password"
            results['status'] = 'AUTH FAILED'
        except paramiko.SSHException as e:
            results['error'] = f"SSH Error: {str(e)}"
            results['status'] = 'SSH ERROR'
        except socket.timeout:
            results['error'] = "Connection timed out"
            results['status'] = 'TIMEOUT'
        except socket.gaierror:
            results['error'] = "Cannot resolve hostname"
            results['status'] = 'DNS ERROR'
        except socket.error as e:
            results['error'] = f"Network Error: {str(e)}"
            results['status'] = 'NET ERROR'
        except Exception as e:
            results['error'] = f"Error: {str(e)}"
            results['status'] = 'ERROR'
        finally:
            try:
                client.close()
            except Exception:
                pass
            self._client = None
            if self.cancelled and results['status'] not in ('DONE',):
                results['status'] = 'CANCELLED'
                if not results['error']:
                    results['error'] = "Execution cancelled"

        return results


# ============== HOST RESULT CARD ==============
class HostResultCard(tk.Frame):
    """Modern result card for each host with beautiful styling"""

    def __init__(self, parent, hostname):
        super().__init__(parent, bg=Theme.BG_DARK, highlightthickness=1,
                        highlightbackground=Theme.BORDER)
        self.hostname = hostname
        self.expanded = False
        self.result_data = None
        self.pulse_active = True
        self.pulse_step = 0
        self.pulse_direction = 1
        self.create_widgets()
        self.start_pulse_smooth()

    def _lerp_color(self, color1, color2, t):
        """Linear interpolate between two hex colors"""
        try:
            r1, g1, b1 = int(color1[1:3], 16), int(color1[3:5], 16), int(color1[5:7], 16)
            r2, g2, b2 = int(color2[1:3], 16), int(color2[3:5], 16), int(color2[5:7], 16)
            r = int(r1 + (r2 - r1) * t)
            g = int(g1 + (g2 - g1) * t)
            b = int(b1 + (b2 - b1) * t)
            return f"#{r:02x}{g:02x}{b:02x}"
        except Exception:
            return color1

    def create_widgets(self):
        # Main container with softer appearance
        main = tk.Frame(self, bg=Theme.BG_CARD)
        main.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)

        # Top accent bar - slightly thinner
        self.accent_bar = tk.Frame(main, bg=Theme.PRIMARY, height=2)
        self.accent_bar.pack(fill=tk.X)

        # Header section with better spacing
        header = tk.Frame(main, bg=Theme.BG_CARD)
        header.pack(fill=tk.X, padx=Theme.PAD_XL, pady=Theme.PAD_LG)

        # Left - Host info with icon
        left = tk.Frame(header, bg=Theme.BG_CARD)
        left.pack(side=tk.LEFT)

        # Status icon with softer appearance
        self.icon_frame = tk.Frame(left, bg=Theme.INFO_BG, width=46, height=46)
        self.icon_frame.pack(side=tk.LEFT)
        self.icon_frame.pack_propagate(False)

        self.icon_label = tk.Label(self.icon_frame, text="...",
                                   font=('Segoe UI', 14, 'bold'),
                                   bg=Theme.INFO_BG, fg=Theme.INFO_LIGHT)
        self.icon_label.place(relx=0.5, rely=0.5, anchor='center')

        # Host details with improved spacing
        info = tk.Frame(left, bg=Theme.BG_CARD)
        info.pack(side=tk.LEFT, padx=Theme.PAD_LG)

        self.host_label = tk.Label(info, text=self.hostname,
                                   font=('Consolas', 13, 'bold'),
                                   bg=Theme.BG_CARD, fg=Theme.TEXT_WHITE)
        self.host_label.pack(anchor=tk.W)

        self.time_label = tk.Label(info, text="Connecting...",
                                   font=('Segoe UI', 9),
                                   bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED)
        self.time_label.pack(anchor=tk.W, pady=(2, 0))

        # Right - Status badge with softer pill shape
        right = tk.Frame(header, bg=Theme.BG_CARD)
        right.pack(side=tk.RIGHT)

        self.status_frame = tk.Frame(right, bg=Theme.INFO_BG)
        self.status_frame.pack()

        self.status_label = tk.Label(self.status_frame, text="  RUNNING  ",
                                     font=('Segoe UI', 10, 'bold'),
                                     bg=Theme.INFO_BG, fg=Theme.INFO_LIGHT)
        self.status_label.pack(padx=Theme.PAD_LG, pady=Theme.PAD_SM)

        # Toggle button with softer styling
        self.toggle_frame = tk.Frame(main, bg=Theme.BG_CARD)
        self.toggle_frame.pack(fill=tk.X, padx=Theme.PAD_XL, pady=Theme.PAD_SM)

        self.toggle_btn = tk.Label(self.toggle_frame,
                                   text="▸  Show Output",
                                   font=('Segoe UI', 10),
                                   bg=Theme.BG_CARD, fg=Theme.CYAN, cursor='hand2')
        self.toggle_btn.pack(anchor=tk.W)
        self.toggle_btn.bind('<Button-1>', self.toggle_output)
        self.toggle_btn.bind('<Enter>', lambda e: self.toggle_btn.config(fg=Theme.CYAN_LIGHT))
        self.toggle_btn.bind('<Leave>', lambda e: self.toggle_btn.config(fg=Theme.CYAN))

        # Output container
        self.output_container = tk.Frame(main, bg=Theme.BG_CARD)

        # Output text area with improved styling
        output_wrapper = tk.Frame(self.output_container, bg=Theme.BG_OUTPUT,
                                 highlightthickness=1, highlightbackground=Theme.BORDER)
        output_wrapper.pack(fill=tk.X, padx=Theme.PAD_XL, pady=Theme.PAD_MD)

        self.output_text = tk.Text(output_wrapper, height=14, font=('Cascadia Code', 10),
                                   bg=Theme.BG_OUTPUT, fg=Theme.TEXT_LIGHT,
                                   relief=tk.FLAT, wrap=tk.WORD,
                                   insertbackground='white', padx=Theme.PAD_LG, pady=Theme.PAD_MD,
                                   selectbackground=Theme.PRIMARY,
                                   selectforeground=Theme.TEXT_WHITE)

        scrollbar = tk.Scrollbar(output_wrapper, command=self.output_text.yview,
                                bg=Theme.BG_CARD, troughcolor=Theme.BG_DARK,
                                activebackground=Theme.PRIMARY)
        self.output_text.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.output_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Configure text tags with vibrant colors
        self.output_text.tag_configure('command', foreground=Theme.CYAN_LIGHT, font=('Cascadia Code', 10, 'bold'))
        self.output_text.tag_configure('output', foreground=Theme.TEXT_LIGHT)
        self.output_text.tag_configure('error', foreground=Theme.ERROR_LIGHT)
        self.output_text.tag_configure('separator', foreground=Theme.TEXT_MUTED)
        self.output_text.tag_configure('success', foreground=Theme.SUCCESS_LIGHT)

        # Error banner with icon
        self.error_banner = tk.Frame(self.output_container, bg=Theme.ERROR_BG)
        self.error_label = tk.Label(self.error_banner, text="",
                                    font=('Segoe UI', 10), bg=Theme.ERROR_BG,
                                    fg=Theme.ERROR_LIGHT, wraplength=550, justify=tk.LEFT)
        self.error_label.pack(padx=20, pady=14, anchor=tk.W)

    def start_pulse_smooth(self):
        """Start smooth pulsing animation while running"""
        if self.pulse_active:
            # Smooth breathing effect between two colors
            max_steps = 20
            t = self.pulse_step / max_steps

            try:
                color = self._lerp_color(Theme.INFO_BG, Theme.INFO, t)
                self.accent_bar.config(bg=color)
            except Exception:
                pass

            self.pulse_step += self.pulse_direction
            if self.pulse_step >= max_steps:
                self.pulse_direction = -1
            elif self.pulse_step <= 0:
                self.pulse_direction = 1

            self.after(50, self.start_pulse_smooth)  # Smoother 20fps animation

    def stop_pulse(self):
        """Stop pulsing animation"""
        self.pulse_active = False

    def toggle_output(self, event=None):
        if self.expanded:
            self.output_container.pack_forget()
            self.toggle_btn.config(text="▸  Show Output")
            self.expanded = False
        else:
            self.output_container.pack(fill=tk.X)
            self.toggle_btn.config(text="▾  Hide Output")
            self.expanded = True

    def get_output_text(self):
        """Get the output text for clipboard copy"""
        return self.output_text.get('1.0', tk.END)

    def update_result(self, result):
        self.result_data = result
        self.stop_pulse()  # Stop pulsing animation

        if result['status'] == 'CANCELLED':
            self.accent_bar.config(bg=Theme.WARNING)
            self.status_frame.config(bg=Theme.WARNING_BG)
            self.status_label.config(text="  CANCELLED  ", bg=Theme.WARNING_BG, fg=Theme.WARNING_LIGHT)
            self.icon_frame.config(bg=Theme.WARNING_BG)
            self.icon_label.config(text="!", bg=Theme.WARNING_BG, fg=Theme.WARNING_LIGHT)
            self.configure(highlightbackground=Theme.WARNING_BORDER, highlightthickness=2)
        elif result['success']:
            self.accent_bar.config(bg=Theme.SUCCESS)
            self.status_frame.config(bg=Theme.SUCCESS_BG)
            self.status_label.config(text="  SUCCESS  ", bg=Theme.SUCCESS_BG, fg=Theme.SUCCESS_GLOW)
            self.icon_frame.config(bg=Theme.SUCCESS_BG)
            self.icon_label.config(text="OK", bg=Theme.SUCCESS_BG, fg=Theme.SUCCESS_GLOW)
            self.configure(highlightbackground=Theme.SUCCESS_BORDER, highlightthickness=2)
        else:
            self.accent_bar.config(bg=Theme.ERROR)
            self.status_frame.config(bg=Theme.ERROR_BG)
            self.status_label.config(text=f"  {result['status']}  ", bg=Theme.ERROR_BG, fg=Theme.ERROR_GLOW)
            self.icon_frame.config(bg=Theme.ERROR_BG)
            self.icon_label.config(text="X", bg=Theme.ERROR_BG, fg=Theme.ERROR_GLOW)
            self.configure(highlightbackground=Theme.ERROR_BORDER, highlightthickness=2)

        self.time_label.config(text=f"Completed: {result['timestamp']}")

        self.output_text.config(state=tk.NORMAL)
        self.output_text.delete('1.0', tk.END)

        if result.get('commands_output') and len(result['commands_output']) > 0:
            total_cmds = len(result['commands_output'])
            self.output_text.insert(tk.END, f"=== EXECUTED {total_cmds} COMMAND(S) ===\n\n", 'separator')

            for i, cmd_data in enumerate(result['commands_output']):
                self.output_text.insert(tk.END, f"[{i+1}/{total_cmds}] ", 'separator')
                self.output_text.insert(tk.END, f"$ {cmd_data['command']}\n", 'command')
                self.output_text.insert(tk.END, "-" * 45 + "\n", 'separator')

                output = cmd_data.get('output', '')
                if output and output != '(no output)':
                    self.output_text.insert(tk.END, f"{output}\n", 'output')
                else:
                    self.output_text.insert(tk.END, "(no output)\n", 'separator')

                error = cmd_data.get('error', '')
                if error:
                    self.output_text.insert(tk.END, f"[STDERR]: {error}\n", 'error')

                if i < total_cmds - 1:
                    self.output_text.insert(tk.END, "\n" + "=" * 45 + "\n\n", 'separator')

        elif result['error']:
            self.output_text.insert(tk.END, f"CONNECTION ERROR:\n{result['error']}\n", 'error')
        else:
            self.output_text.insert(tk.END, "(No commands executed)\n", 'output')

        self.output_text.config(state=tk.DISABLED)

        if result['error'] and not result['success']:
            self.error_label.config(text=f"ERROR: {result['error']}")
            self.error_banner.pack(fill=tk.X, padx=16, pady=8)

        if not result['success'] and not self.expanded:
            self.toggle_output()


# ============== MAIN APPLICATION ==============
class MultiHostExecutorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("MultiHost SSH")
        self.root.geometry("1350x900")
        self.root.minsize(1100, 700)
        self.root.configure(bg=Theme.BG_MAIN)

        # Set application icon
        self._set_app_icon()

        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

        self.result_queue = queue.Queue()
        self.is_executing = False
        self.cancel_requested = False
        self.host_cards = {}
        self.all_results = []
        self.executors = []
        self.preset_manager = CommandPresetManager()
        self.host_group_manager = HostGroupManager()
        self.session_manager = SessionConfigManager()
        self.auth_manager = AdminAuthManager()
        self.admin_unlocked = False

        # Lock the source file as read-only by default — only admin can unlock it
        if self.auth_manager.is_configured():
            try:
                self._set_source_readonly(True)
            except Exception:
                pass

        self.setup_styles()
        self.create_widgets()
        self.load_session()
        self.setup_keyboard_shortcuts()
        self._attach_tooltips()
        self.check_queue()

    def _attach_tooltips(self):
        """Attach helpful tooltips to controls so users know what each does."""
        tips = [
            (self.user_entry, "Linux SSH username (same on every host)."),
            (self.pass_entry, "SSH password. Used for sudo too if a sudo command is in your list."),
            (self.key_file_entry, "Path to your private SSH key (RSA / Ed25519 / ECDSA / DSA)."),
            (self.key_passphrase_entry, "Leave empty if your key isn't encrypted."),
            (self.port_entry, "SSH port. Default is 22."),
            (self.timeout_entry, "Per-host connection timeout in seconds."),
            (self.max_concurrent_entry, "How many hosts to connect to in parallel (1–100)."),
            (self.hosts_text, "One host per line. # comments are ignored. Use Import to load a file."),
            (self.commands_text, "One command per line. Sudo passwords are injected automatically."),
            (self.execute_btn, "Run all commands on all hosts (Ctrl+Enter)."),
            (self.stop_btn, "Stop in-flight execution (Esc)."),
            (self.rerun_btn, "Re-run only on hosts that failed last time."),
            (self.clear_btn, "Clear results and reset progress."),
            (self.save_session_btn, "Remember username, hosts, commands, and key path for next launch."),
            (self.copy_btn, "Copy all results to clipboard."),
            (self.excel_btn, "Export results to an Excel (.xlsx) file."),
            (self.export_btn, "Export results to a plain text file."),
            (self.broadcast_entry, "One-shot command sent to already-connected hosts."),
            (self.broadcast_btn, "Send the broadcast command to selected target hosts."),
            (self.admin_btn, "Admin lock: enter password to enable code editing."),
        ]
        for widget, text in tips:
            if widget is not None:
                Tooltip(widget, text)

    def setup_keyboard_shortcuts(self):
        """Setup keyboard shortcuts for quick actions"""
        self.root.bind('<Control-Return>', lambda e: self.execute_commands())
        self.root.bind('<Control-e>', lambda e: self.execute_commands())
        self.root.bind('<Escape>', lambda e: self.stop_execution())
        self.root.bind('<Control-l>', lambda e: self.clear_results())

    def _set_app_icon(self):
        """Set the application icon for taskbar and window"""
        try:
            # Create icons at multiple sizes for better Windows taskbar display
            icon_sizes = [64, 32, 16]
            icons = []

            for size in icon_sizes:
                icon = tk.PhotoImage(width=size, height=size)

                # Calculate proportions based on icon size
                border = max(1, size // 16)
                title_height = size // 4
                sep_y = title_height + border

                # Build icon row by row for better performance
                for y in range(size):
                    row_colors = []
                    for x in range(size):
                        if border <= x < size - border and border <= y < size - border:
                            if y < title_height:
                                # Title bar - purple gradient
                                row_colors.append("#764ba2")
                            elif y == sep_y:
                                # Separator
                                row_colors.append("#4a3f6b")
                            else:
                                # Terminal body
                                inner_x = x - border
                                inner_y = y - sep_y - 1
                                body_width = size - 2 * border
                                body_height = size - sep_y - border - 1

                                # Add terminal "text" lines
                                line_height = max(1, body_height // 8)
                                text_start = max(2, size // 10)

                                if body_height > 0:
                                    line_pos = inner_y / body_height if body_height > 0 else 0
                                    if 0.15 <= line_pos < 0.25 and text_start <= inner_x < body_width * 0.7:
                                        row_colors.append("#00d4aa")  # Cyan text
                                    elif 0.35 <= line_pos < 0.45 and text_start <= inner_x < body_width * 0.5:
                                        row_colors.append("#667eea")  # Blue text
                                    elif 0.55 <= line_pos < 0.65 and text_start <= inner_x < body_width * 0.85:
                                        row_colors.append("#f093fb")  # Pink text
                                    elif 0.75 <= line_pos < 0.85 and text_start <= inner_x < text_start + max(3, size // 8):
                                        row_colors.append("#00ff88")  # Green cursor
                                    else:
                                        row_colors.append("#1a1a2e")  # Dark background
                                else:
                                    row_colors.append("#1a1a2e")
                        elif 0 <= x < size and 0 <= y < size:
                            # Rounded border effect
                            row_colors.append("#667eea")
                        else:
                            row_colors.append("#16213e")

                    # Put entire row at once for performance
                    icon.put("{" + " ".join(row_colors) + "}", (0, y))

                icons.append(icon)

            # Set icon with multiple sizes (Windows uses appropriate size)
            self.root.iconphoto(True, *icons)
            # Keep references to prevent garbage collection
            self._icon_images = icons

        except Exception:
            pass  # If all else fails, use default icon

    def setup_styles(self):
        """Configure ttk styles for modern look"""
        style = ttk.Style()
        style.theme_use('clam')

        # Combobox style
        style.configure('Modern.TCombobox',
                       fieldbackground=Theme.BG_INPUT,
                       background=Theme.BG_INPUT,
                       foreground=Theme.TEXT_WHITE,
                       arrowcolor=Theme.TEXT_WHITE,
                       bordercolor=Theme.BORDER,
                       lightcolor=Theme.BORDER,
                       darkcolor=Theme.BORDER)

        style.map('Modern.TCombobox',
                 fieldbackground=[('readonly', Theme.BG_INPUT)],
                 selectbackground=[('readonly', Theme.PRIMARY)],
                 selectforeground=[('readonly', Theme.TEXT_WHITE)])

        # Scrollbar style
        style.configure('Modern.Vertical.TScrollbar',
                       background=Theme.BG_CARD,
                       troughcolor=Theme.BG_DARK,
                       bordercolor=Theme.BG_DARK,
                       arrowcolor=Theme.TEXT_MEDIUM)

    def create_widgets(self):
        # ====== STUNNING HEADER ======
        self.header = StunningHeader(self.root)

        # ====== MAIN CONTENT ======
        content = tk.Frame(self.root, bg=Theme.BG_MAIN)
        content.pack(fill=tk.BOTH, expand=True, padx=16, pady=10)

        # ====== LEFT PANEL (Scrollable) ======
        left_container = tk.Frame(content, bg=Theme.BG_MAIN, width=450)
        left_container.pack(side=tk.LEFT, fill=tk.Y, padx=8)
        left_container.pack_propagate(False)

        # Create canvas for scrolling
        left_canvas = tk.Canvas(left_container, bg=Theme.BG_MAIN, highlightthickness=0, width=430)
        left_scrollbar = ttk.Scrollbar(left_container, orient=tk.VERTICAL, command=left_canvas.yview,
                                       style='Modern.Vertical.TScrollbar')
        left = tk.Frame(left_canvas, bg=Theme.BG_MAIN)

        left_canvas.configure(yscrollcommand=left_scrollbar.set)
        left_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        left_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        left_canvas.create_window((0, 0), window=left, anchor=tk.NW, width=420)
        left.bind('<Configure>', lambda e: left_canvas.configure(scrollregion=left_canvas.bbox("all")))

        # Mouse wheel scrolling
        def _on_mousewheel(event):
            left_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        left_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # --- TARGET HOSTS --- (Cyan colored section)
        ColorfulSectionHeader(left, "TARGET HOSTS", "Server addresses to connect",
                             icon="H", accent_color=Theme.SECTION_HOSTS)

        hosts_card = self.create_card(left, accent_color=Theme.SECTION_HOSTS)

        # Host groups row
        group_row = tk.Frame(hosts_card, bg=Theme.BG_CARD)
        group_row.pack(fill=tk.X, padx=12, pady=5)

        tk.Label(group_row, text="Host Groups:", font=('Segoe UI', 9, 'bold'),
                bg=Theme.BG_CARD, fg=Theme.TEXT_MEDIUM).pack(side=tk.LEFT)

        self.host_group_var = tk.StringVar()
        self.host_group_combo = ttk.Combobox(group_row, textvariable=self.host_group_var,
                                              style='Modern.TCombobox',
                                              state='readonly', width=12, font=('Segoe UI', 9))
        self.host_group_combo.pack(side=tk.LEFT, padx=8)
        self.update_host_group_list()

        btn_style = {'font': ('Segoe UI', 8, 'bold'), 'relief': tk.FLAT,
                    'cursor': 'hand2', 'width': 5, 'fg': Theme.TEXT_WHITE}

        tk.Button(group_row, text="Load", bg=Theme.INFO,
                 command=self.load_host_group, **btn_style).pack(side=tk.LEFT, padx=2, ipady=2)
        tk.Button(group_row, text="Save", bg=Theme.SUCCESS,
                 command=self.save_host_group, **btn_style).pack(side=tk.LEFT, padx=2, ipady=2)
        tk.Button(group_row, text="Del", bg=Theme.ERROR,
                 command=self.delete_host_group, **btn_style).pack(side=tk.LEFT, padx=2, ipady=2)
        tk.Button(group_row, text="Import", bg=Theme.SECONDARY,
                 command=self.import_hosts, **btn_style).pack(side=tk.LEFT, padx=2, ipady=2)

        tk.Label(hosts_card, text="One hostname or IP per line. Lines starting with # are ignored.",
                font=('Segoe UI', 9), bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor=tk.W, padx=12, pady=2)

        hosts_frame = tk.Frame(hosts_card, bg=Theme.BG_INPUT, highlightbackground=Theme.BORDER, highlightthickness=1)
        hosts_frame.pack(fill=tk.X, padx=12, pady=5)

        self.hosts_text = tk.Text(hosts_frame, height=5, font=('Consolas', 11),
                                  bg=Theme.BG_INPUT, fg=Theme.TEXT_WHITE, relief=tk.FLAT,
                                  insertbackground=Theme.TEXT_WHITE, selectbackground=Theme.PRIMARY)
        self.hosts_text.pack(fill=tk.X, padx=6, pady=5)
        self.hosts_text.insert(tk.END, "192.168.1.10\n192.168.1.11")

        # --- CREDENTIALS --- (Amber colored section)
        ColorfulSectionHeader(left, "CREDENTIALS", "SSH authentication details",
                             icon="K", accent_color=Theme.SECTION_CREDS)

        cred_card = self.create_card(left, accent_color=Theme.SECTION_CREDS)

        # Authentication method
        auth_row = tk.Frame(cred_card, bg=Theme.BG_CARD)
        auth_row.pack(fill=tk.X, padx=12, pady=4)

        tk.Label(auth_row, text="Auth Method:", font=('Segoe UI', 9, 'bold'),
                bg=Theme.BG_CARD, fg=Theme.TEXT_MEDIUM).pack(side=tk.LEFT)

        self.auth_method_var = tk.StringVar(value="password")
        tk.Radiobutton(auth_row, text="Password", variable=self.auth_method_var, value="password",
                      bg=Theme.BG_CARD, fg=Theme.TEXT_LIGHT, selectcolor=Theme.BG_INPUT,
                      activebackground=Theme.BG_CARD, activeforeground=Theme.TEXT_WHITE,
                      font=('Segoe UI', 9), command=self.toggle_auth_method).pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(auth_row, text="SSH Key", variable=self.auth_method_var, value="key",
                      bg=Theme.BG_CARD, fg=Theme.TEXT_LIGHT, selectcolor=Theme.BG_INPUT,
                      activebackground=Theme.BG_CARD, activeforeground=Theme.TEXT_WHITE,
                      font=('Segoe UI', 9), command=self.toggle_auth_method).pack(side=tk.LEFT)

        # Username
        self.create_input_field(cred_card, "Username")
        user_frame = tk.Frame(cred_card, bg=Theme.BG_INPUT, highlightbackground=Theme.BORDER, highlightthickness=1)
        user_frame.pack(fill=tk.X, padx=12, pady=3)
        self.user_entry = tk.Entry(user_frame, font=('Consolas', 11),
                                   bg=Theme.BG_INPUT, fg=Theme.TEXT_WHITE, relief=tk.FLAT,
                                   insertbackground=Theme.TEXT_WHITE)
        self.user_entry.pack(fill=tk.X, padx=6, pady=5)

        # Password frame
        self.password_frame = tk.Frame(cred_card, bg=Theme.BG_CARD)
        self.password_frame.pack(fill=tk.X)

        self.create_input_field(self.password_frame, "Password")
        pass_frame = tk.Frame(self.password_frame, bg=Theme.BG_INPUT, highlightbackground=Theme.BORDER, highlightthickness=1)
        pass_frame.pack(fill=tk.X, padx=12, pady=3)
        self.pass_entry = tk.Entry(pass_frame, font=('Consolas', 11),
                                   bg=Theme.BG_INPUT, fg=Theme.TEXT_WHITE, relief=tk.FLAT,
                                   insertbackground=Theme.TEXT_WHITE, show="*")
        self.pass_entry.pack(fill=tk.X, padx=6, pady=5)

        # Show password checkbox
        self.show_pass_var = tk.BooleanVar()
        check_frame = tk.Frame(self.password_frame, bg=Theme.BG_CARD)
        check_frame.pack(fill=tk.X, padx=12, pady=2)
        tk.Checkbutton(check_frame, text="Show password", variable=self.show_pass_var,
                      command=self.toggle_password, bg=Theme.BG_CARD,
                      fg=Theme.TEXT_MEDIUM, font=('Segoe UI', 9),
                      selectcolor=Theme.BG_INPUT, activebackground=Theme.BG_CARD,
                      activeforeground=Theme.TEXT_WHITE).pack(anchor=tk.W)

        # SSH Key frame (hidden by default)
        self.key_frame = tk.Frame(cred_card, bg=Theme.BG_CARD)

        self.create_input_field(self.key_frame, "SSH Key File")
        key_input_frame = tk.Frame(self.key_frame, bg=Theme.BG_CARD)
        key_input_frame.pack(fill=tk.X, padx=14, pady=4)

        key_entry_frame = tk.Frame(key_input_frame, bg=Theme.BG_INPUT, highlightbackground=Theme.BORDER, highlightthickness=1)
        key_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.key_file_entry = tk.Entry(key_entry_frame, font=('Consolas', 10),
                                       bg=Theme.BG_INPUT, fg=Theme.TEXT_WHITE, relief=tk.FLAT,
                                       insertbackground=Theme.TEXT_WHITE)
        self.key_file_entry.pack(fill=tk.X, padx=8, pady=6)

        tk.Button(key_input_frame, text="Browse", bg=Theme.INFO, fg=Theme.TEXT_WHITE,
                 font=('Segoe UI', 8, 'bold'), relief=tk.FLAT, cursor='hand2',
                 command=self.browse_key_file).pack(side=tk.LEFT, padx=8, ipady=4)

        self.create_input_field(self.key_frame, "Key Passphrase (optional)")
        key_pass_frame = tk.Frame(self.key_frame, bg=Theme.BG_INPUT, highlightbackground=Theme.BORDER, highlightthickness=1)
        key_pass_frame.pack(fill=tk.X, padx=14, pady=4)
        self.key_passphrase_entry = tk.Entry(key_pass_frame, font=('Consolas', 11),
                                             bg=Theme.BG_INPUT, fg=Theme.TEXT_WHITE, relief=tk.FLAT,
                                             insertbackground=Theme.TEXT_WHITE, show="*")
        self.key_passphrase_entry.pack(fill=tk.X, padx=8, pady=8)

        # Port, Timeout, and Max Concurrent row
        row = tk.Frame(cred_card, bg=Theme.BG_CARD)
        row.pack(fill=tk.X, padx=12, pady=4)

        # Port
        port_col = tk.Frame(row, bg=Theme.BG_CARD)
        port_col.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
        tk.Label(port_col, text="Port", font=('Segoe UI', 9),
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor=tk.W)
        port_input = tk.Frame(port_col, bg=Theme.BG_INPUT, highlightbackground=Theme.BORDER, highlightthickness=1)
        port_input.pack(fill=tk.X, pady=4)
        self.port_entry = tk.Entry(port_input, font=('Consolas', 11), width=8,
                                   bg=Theme.BG_INPUT, fg=Theme.TEXT_WHITE, relief=tk.FLAT,
                                   insertbackground=Theme.TEXT_WHITE)
        self.port_entry.pack(padx=8, pady=6)
        self.port_entry.insert(0, "22")

        # Timeout
        timeout_col = tk.Frame(row, bg=Theme.BG_CARD)
        timeout_col.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
        tk.Label(timeout_col, text="Timeout", font=('Segoe UI', 9),
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor=tk.W)
        timeout_input = tk.Frame(timeout_col, bg=Theme.BG_INPUT, highlightbackground=Theme.BORDER, highlightthickness=1)
        timeout_input.pack(fill=tk.X, pady=4)
        self.timeout_entry = tk.Entry(timeout_input, font=('Consolas', 11), width=8,
                                      bg=Theme.BG_INPUT, fg=Theme.TEXT_WHITE, relief=tk.FLAT,
                                      insertbackground=Theme.TEXT_WHITE)
        self.timeout_entry.pack(padx=8, pady=6)
        self.timeout_entry.insert(0, "30")

        # Max Concurrent
        max_col = tk.Frame(row, bg=Theme.BG_CARD)
        max_col.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
        tk.Label(max_col, text="Max Concurrent", font=('Segoe UI', 9),
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor=tk.W)
        max_input = tk.Frame(max_col, bg=Theme.BG_INPUT, highlightbackground=Theme.BORDER, highlightthickness=1)
        max_input.pack(fill=tk.X, pady=4)
        self.max_concurrent_entry = tk.Entry(max_input, font=('Consolas', 11), width=8,
                                             bg=Theme.BG_INPUT, fg=Theme.TEXT_WHITE, relief=tk.FLAT,
                                             insertbackground=Theme.TEXT_WHITE)
        self.max_concurrent_entry.pack(padx=8, pady=6)
        self.max_concurrent_entry.insert(0, "10")

        # --- COMMANDS --- (Purple colored section)
        ColorfulSectionHeader(left, "COMMANDS", "Commands to execute on all hosts",
                             icon="$", accent_color=Theme.SECTION_CMDS)

        cmd_card = self.create_card(left, accent_color=Theme.SECTION_CMDS)

        # Preset row
        preset_row = tk.Frame(cmd_card, bg=Theme.BG_CARD)
        preset_row.pack(fill=tk.X, padx=12, pady=5)

        tk.Label(preset_row, text="Presets:", font=('Segoe UI', 9, 'bold'),
                bg=Theme.BG_CARD, fg=Theme.TEXT_MEDIUM).pack(side=tk.LEFT)

        self.preset_var = tk.StringVar()
        self.preset_combo = ttk.Combobox(preset_row, textvariable=self.preset_var,
                                         style='Modern.TCombobox',
                                         state='readonly', width=14, font=('Segoe UI', 9))
        self.preset_combo.pack(side=tk.LEFT, padx=8)
        self.update_preset_list()
        self.preset_combo.bind('<<ComboboxSelected>>', self.on_preset_selected)

        # Preset buttons
        btn_style = {'font': ('Segoe UI', 8, 'bold'), 'relief': tk.FLAT,
                    'cursor': 'hand2', 'width': 6, 'fg': Theme.TEXT_WHITE}

        tk.Button(preset_row, text="Load", bg=Theme.INFO,
                 command=self.load_preset, **btn_style).pack(side=tk.LEFT, padx=2, ipady=2)
        tk.Button(preset_row, text="Save", bg=Theme.SUCCESS,
                 command=self.save_preset, **btn_style).pack(side=tk.LEFT, padx=2, ipady=2)
        tk.Button(preset_row, text="Del", bg=Theme.ERROR,
                 command=self.delete_preset, **btn_style).pack(side=tk.LEFT, padx=2, ipady=2)

        tk.Label(cmd_card, text="One command per line. Sudo passwords are handled automatically.",
                font=('Segoe UI', 9), bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor=tk.W, padx=12, pady=2)

        cmd_frame = tk.Frame(cmd_card, bg=Theme.BG_INPUT, highlightbackground=Theme.BORDER, highlightthickness=1)
        cmd_frame.pack(fill=tk.X, padx=12, pady=5)

        self.commands_text = tk.Text(cmd_frame, height=4, font=('Consolas', 11),
                                     bg=Theme.BG_INPUT, fg=Theme.TEXT_WHITE, relief=tk.FLAT,
                                     insertbackground=Theme.TEXT_WHITE, selectbackground=Theme.PRIMARY)
        self.commands_text.pack(fill=tk.X, padx=6, pady=5)
        self.commands_text.insert(tk.END, "hostname\nuptime\ndf -h")

        # === ACTION BUTTONS INSIDE COMMAND CARD ===
        # Separator line
        tk.Frame(cmd_card, bg=Theme.SECTION_CMDS, height=2).pack(fill=tk.X, padx=12, pady=4)

        # EXECUTE Button - BIG and GREEN
        self.execute_btn = tk.Button(cmd_card, text="▶  EXECUTE COMMANDS",
                                     font=('Segoe UI', 12, 'bold'),
                                     bg="#10b981", fg="#ffffff",
                                     activebackground="#34d399",
                                     activeforeground="#ffffff",
                                     relief=tk.FLAT, cursor='hand2',
                                     command=self.execute_commands)
        self.execute_btn.pack(fill=tk.X, padx=12, pady=5, ipady=6)

        # Control buttons row - unified sizing for consistent alignment
        ctrl_frame = tk.Frame(cmd_card, bg=Theme.BG_CARD)
        ctrl_frame.pack(fill=tk.X, padx=12, pady=4)

        ctrl_btn_style = {
            'font': ('Segoe UI', 9, 'bold'),
            'fg': '#ffffff',
            'relief': tk.FLAT,
            'cursor': 'hand2',
            'width': 7,
            'bd': 0,
            'highlightthickness': 0,
        }
        ctrl_pack = {'side': tk.LEFT, 'padx': 3, 'ipady': 4}
        ctrl_pack_right = {'side': tk.RIGHT, 'padx': 3, 'ipady': 4}

        self.stop_btn = tk.Button(ctrl_frame, text="STOP",
                                  bg="#ef4444", activebackground="#f87171",
                                  state='disabled',
                                  command=self.stop_execution, **ctrl_btn_style)
        self.stop_btn.pack(**ctrl_pack)

        self.rerun_btn = tk.Button(ctrl_frame, text="RETRY",
                                   bg="#f59e0b", activebackground="#fbbf24",
                                   command=self.rerun_failed, **ctrl_btn_style)
        self.rerun_btn.pack(**ctrl_pack)

        self.clear_btn = tk.Button(ctrl_frame, text="CLEAR",
                                   bg="#64748b", activebackground="#94a3b8",
                                   command=self.clear_results, **ctrl_btn_style)
        self.clear_btn.pack(**ctrl_pack)

        # Export buttons on right - same dimensions as left buttons
        self.save_session_btn = tk.Button(ctrl_frame, text="SAVE",
                                          bg="#ec4899", activebackground="#f472b6",
                                          command=self.save_session, **ctrl_btn_style)
        self.save_session_btn.pack(**ctrl_pack_right)

        self.copy_btn = tk.Button(ctrl_frame, text="COPY",
                                  bg="#8b5cf6", activebackground="#a78bfa",
                                  command=self.copy_results_to_clipboard, **ctrl_btn_style)
        self.copy_btn.pack(**ctrl_pack_right)

        self.excel_btn = tk.Button(ctrl_frame, text="EXCEL",
                                   bg="#84cc16", activebackground="#a3e635",
                                   command=self.export_to_excel, **ctrl_btn_style)
        self.excel_btn.pack(**ctrl_pack_right)

        self.export_btn = tk.Button(ctrl_frame, text="TXT",
                                    bg="#14b8a6", activebackground="#2dd4bf",
                                    command=self.export_results, **ctrl_btn_style)
        self.export_btn.pack(**ctrl_pack_right)

        # Bottom padding
        tk.Frame(cmd_card, bg=Theme.BG_CARD, height=4).pack(fill=tk.X)

        # ====== RIGHT PANEL - RESULTS ======
        right = tk.Frame(content, bg=Theme.BG_MAIN)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

        # Results header with gradient accent
        results_header = tk.Frame(right, bg=Theme.BG_MAIN)
        results_header.pack(fill=tk.X, pady=8)

        # Left side: Icon and title
        header_left = tk.Frame(results_header, bg=Theme.BG_MAIN)
        header_left.pack(side=tk.LEFT)

        # Results icon
        results_icon = tk.Frame(header_left, bg=Theme.SECTION_RESULTS, width=36, height=36)
        results_icon.pack(side=tk.LEFT, padx=(0, 10))
        results_icon.pack_propagate(False)
        tk.Label(results_icon, text="R", font=('Segoe UI', 14, 'bold'),
                bg=Theme.SECTION_RESULTS, fg=Theme.TEXT_WHITE).place(relx=0.5, rely=0.5, anchor='center')

        tk.Label(header_left, text="EXECUTION RESULTS",
                font=('Segoe UI', 14, 'bold'),
                bg=Theme.BG_MAIN, fg=Theme.SECTION_RESULTS).pack(side=tk.LEFT)

        # Stats badges with vibrant colors
        stats = tk.Frame(results_header, bg=Theme.BG_MAIN)
        stats.pack(side=tk.RIGHT)

        # Total badge
        total_bg = tk.Frame(stats, bg=Theme.PURPLE)
        total_bg.pack(side=tk.LEFT, padx=4)
        self.total_label = tk.Label(total_bg, text=" TOTAL: 0 ", font=('Segoe UI', 10, 'bold'),
                                   bg=Theme.PURPLE, fg=Theme.TEXT_WHITE)
        self.total_label.pack(padx=12, pady=6)

        # Success badge
        done_bg = tk.Frame(stats, bg=Theme.SUCCESS)
        done_bg.pack(side=tk.LEFT, padx=4)
        self.done_label = tk.Label(done_bg, text=" SUCCESS: 0 ", font=('Segoe UI', 10, 'bold'),
                                  bg=Theme.SUCCESS, fg=Theme.TEXT_WHITE)
        self.done_label.pack(padx=12, pady=6)

        # Failed badge
        fail_bg = tk.Frame(stats, bg=Theme.ERROR)
        fail_bg.pack(side=tk.LEFT, padx=4)
        self.failed_label = tk.Label(fail_bg, text=" FAILED: 0 ", font=('Segoe UI', 10, 'bold'),
                                    bg=Theme.ERROR, fg=Theme.TEXT_WHITE)
        self.failed_label.pack(padx=12, pady=6)

        # Animated progress bar with gradient effect
        progress_frame = tk.Frame(right, bg=Theme.BG_MAIN)
        progress_frame.pack(fill=tk.X, pady=8)

        progress_bg = tk.Frame(progress_frame, bg=Theme.BG_DARK, height=10)
        progress_bg.pack(fill=tk.X)
        progress_bg.pack_propagate(False)

        self.progress_bar = tk.Frame(progress_bg, bg=Theme.INDIGO, width=0)
        self.progress_bar.pack(side=tk.LEFT, fill=tk.Y)

        self.status_label = tk.Label(progress_frame, text="Ready to execute",
                                    font=('Segoe UI', 11),
                                    bg=Theme.BG_MAIN, fg=Theme.CYAN)
        self.status_label.pack(anchor=tk.W, pady=8)

        # --- DOWNLOAD RESULTS SECTION (hidden initially) ---
        self.download_frame = tk.Frame(right, bg=Theme.BG_MAIN)
        # Don't pack yet - will be shown after execution

        self.download_btn = tk.Button(self.download_frame,
                                      text="⬇  DOWNLOAD RESULTS (.txt)",
                                      font=('Segoe UI', 12, 'bold'),
                                      bg=Theme.GOLD, fg="#1a1500",
                                      activebackground="#fcd34d",
                                      activeforeground="#1a1500",
                                      relief=tk.FLAT, cursor='hand2',
                                      command=self.download_results)
        self.download_btn.pack(fill=tk.X, pady=5, ipady=12)

        self.download_sublabel = tk.Label(self.download_frame,
                                          text="Save execution log as text file",
                                          font=('Segoe UI', 9),
                                          bg=Theme.BG_MAIN, fg=Theme.TEXT_MUTED)
        self.download_sublabel.pack(anchor=tk.CENTER)

        # Broadcast Input Frame with colorful border
        self.broadcast_outer = tk.Frame(right, bg=Theme.PINK)
        self.broadcast_outer.pack(fill=tk.X, pady=8)

        broadcast_frame = tk.Frame(self.broadcast_outer, bg=Theme.BG_CARD)
        broadcast_frame.pack(fill=tk.BOTH, padx=2, pady=2)

        # Broadcast icon
        broadcast_icon = tk.Frame(broadcast_frame, bg=Theme.PINK, width=32, height=32)
        broadcast_icon.pack(side=tk.LEFT, padx=12, pady=10)
        broadcast_icon.pack_propagate(False)
        tk.Label(broadcast_icon, text="B", font=('Segoe UI', 12, 'bold'),
                bg=Theme.PINK, fg=Theme.TEXT_WHITE).place(relx=0.5, rely=0.5, anchor='center')

        tk.Label(broadcast_frame, text="BROADCAST", font=('Segoe UI', 10, 'bold'),
                bg=Theme.BG_CARD, fg=Theme.PINK).pack(side=tk.LEFT, pady=10)

        tk.Label(broadcast_frame, text="Send to all:", font=('Segoe UI', 9),
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(side=tk.LEFT, padx=(10, 5))

        broadcast_entry_frame = tk.Frame(broadcast_frame, bg=Theme.BG_INPUT,
                                        highlightbackground=Theme.BORDER, highlightthickness=1)
        broadcast_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=8)

        self.broadcast_entry = tk.Entry(broadcast_entry_frame, font=('Consolas', 11),
                                        bg=Theme.BG_INPUT, fg=Theme.TEXT_WHITE, relief=tk.FLAT,
                                        insertbackground=Theme.TEXT_WHITE)
        self.broadcast_entry.pack(fill=tk.X, padx=8, pady=6)
        self.broadcast_entry.bind('<Return>', lambda e: self.broadcast_command())

        self.broadcast_btn = tk.Button(broadcast_frame, text="SEND TO ALL", bg=Theme.PINK,
                                       fg=Theme.TEXT_WHITE, font=('Segoe UI', 9, 'bold'),
                                       relief=tk.FLAT, cursor='hand2', padx=15,
                                       activebackground=Theme.PINK_LIGHT,
                                       activeforeground=Theme.TEXT_WHITE,
                                       command=self.broadcast_command)
        self.broadcast_btn.pack(side=tk.LEFT, padx=10, pady=8, ipady=5)

        self.broadcast_success_only_var = tk.BooleanVar(value=True)
        tk.Checkbutton(broadcast_frame, text="Success only", variable=self.broadcast_success_only_var,
                      bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED, font=('Segoe UI', 8),
                      selectcolor=Theme.BG_INPUT, activebackground=Theme.BG_CARD,
                      activeforeground=Theme.TEXT_WHITE).pack(side=tk.LEFT, padx=5)

        # Results container
        results_container = tk.Frame(right, bg=Theme.BG_DARK)
        results_container.pack(fill=tk.BOTH, expand=True)

        self.results_canvas = tk.Canvas(results_container, bg=Theme.BG_DARK, highlightthickness=0)
        scrollbar = ttk.Scrollbar(results_container, orient=tk.VERTICAL,
                                 command=self.results_canvas.yview,
                                 style='Modern.Vertical.TScrollbar')

        self.results_frame = tk.Frame(self.results_canvas, bg=Theme.BG_DARK)

        self.results_canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.results_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.canvas_window = self.results_canvas.create_window((0, 0), window=self.results_frame, anchor=tk.NW)

        self.results_frame.bind('<Configure>',
            lambda e: self.results_canvas.configure(scrollregion=self.results_canvas.bbox("all")))
        self.results_canvas.bind('<Configure>',
            lambda e: self.results_canvas.itemconfig(self.canvas_window, width=e.width))
        self.results_canvas.bind_all('<MouseWheel>',
            lambda e: self.results_canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # Empty state — friendly onboarding guidance
        self._build_empty_state()

        # ====== STUNNING FOOTER ======
        footer_container = tk.Frame(self.root, bg=Theme.BG_DARK)
        footer_container.pack(fill=tk.X, side=tk.BOTTOM)

        # Animated color bar at top of footer
        self.footer_glow = tk.Frame(footer_container, height=3, bg=Theme.PRIMARY)
        self.footer_glow.pack(fill=tk.X)

        footer = tk.Frame(footer_container, bg="#0f0f1a", height=40)
        footer.pack(fill=tk.X)
        footer.pack_propagate(False)

        # Left side - Keyboard shortcuts hint
        left_hint = tk.Frame(footer, bg="#0f0f1a")
        left_hint.pack(side=tk.LEFT, padx=20, pady=10)
        tk.Label(left_hint, text="Ctrl+Enter", font=('Consolas', 8),
                bg="#0f0f1a", fg=Theme.CYAN).pack(side=tk.LEFT)
        tk.Label(left_hint, text=" Execute", font=('Segoe UI', 8),
                bg="#0f0f1a", fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)

        # Center - Copyright with colorful text
        footer_content = tk.Frame(footer, bg="#0f0f1a")
        footer_content.place(relx=0.5, rely=0.5, anchor='center')

        tk.Label(footer_content, text="© 2026 ", font=('Segoe UI', 9),
                bg="#0f0f1a", fg=Theme.CYAN).pack(side=tk.LEFT)
        tk.Label(footer_content, text="MultiHost SSH", font=('Segoe UI', 9, 'bold'),
                bg="#0f0f1a", fg=Theme.TEXT_WHITE).pack(side=tk.LEFT)
        tk.Label(footer_content, text="  •  Crafted with ", font=('Segoe UI', 9),
                bg="#0f0f1a", fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)
        tk.Label(footer_content, text="♥", font=('Segoe UI', 10),
                bg="#0f0f1a", fg=Theme.ERROR).pack(side=tk.LEFT)
        tk.Label(footer_content, text=" by ", font=('Segoe UI', 9),
                bg="#0f0f1a", fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)
        tk.Label(footer_content, text="SURYA SSK", font=('Segoe UI', 9, 'bold'),
                bg="#0f0f1a", fg=Theme.PINK).pack(side=tk.LEFT)
        tk.Label(footer_content, text="  •  ", font=('Segoe UI', 9),
                bg="#0f0f1a", fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)

        email_link = tk.Label(footer_content, text="sl.suryassk@outlook.com",
                              font=('Segoe UI', 9, 'underline'),
                              bg="#0f0f1a", fg=Theme.CYAN_LIGHT, cursor='hand2')
        email_link.pack(side=tk.LEFT)
        email_link.bind('<Button-1>',
                        lambda e: __import__('webbrowser').open(
                            'mailto:sl.suryassk@outlook.com?subject=MultiHost%20Executor'))
        email_link.bind('<Enter>', lambda e: email_link.config(fg=Theme.PINK))
        email_link.bind('<Leave>', lambda e: email_link.config(fg=Theme.CYAN_LIGHT))

        # Right side - Version + Admin lock
        right_info = tk.Frame(footer, bg="#0f0f1a")
        right_info.pack(side=tk.RIGHT, padx=20, pady=8)

        tk.Label(right_info, text="v2.0", font=('Consolas', 9, 'bold'),
                bg="#0f0f1a", fg=Theme.SUCCESS).pack(side=tk.RIGHT, padx=(10, 0))

        self.admin_btn = tk.Button(right_info, text="🔒 Admin",
                                   font=('Segoe UI', 9, 'bold'),
                                   bg="#475569", fg="#ffffff",
                                   activebackground="#64748b",
                                   relief=tk.FLAT, cursor='hand2', bd=0,
                                   command=self.toggle_admin_mode)
        self.admin_btn.pack(side=tk.RIGHT, ipadx=8, ipady=2)

        # Start footer animation
        self.animate_footer_glow()

    def _on_execute_hover(self, event):
        """Hover effect for execute button"""
        hover_color = Theme.SUCCESS_LIGHT
        for widget in [self.big_execute_frame, self.execute_icon, self.execute_label,
                       self.execute_sublabel, self.execute_hint]:
            try:
                if widget.winfo_exists():
                    widget.config(bg=hover_color)
            except Exception:
                pass

    def _on_execute_leave(self, event):
        """Leave effect for execute button"""
        normal_color = Theme.SUCCESS
        for widget in [self.big_execute_frame, self.execute_icon, self.execute_label]:
            try:
                if widget.winfo_exists():
                    widget.config(bg=normal_color)
            except Exception:
                pass
        try:
            self.execute_sublabel.config(bg=normal_color)
            self.execute_hint.config(bg=normal_color)
        except Exception:
            pass

    def download_results(self):
        """Download results as text file directly to Downloads folder"""
        if not self.all_results:
            messagebox.showinfo("Info", "No results to download")
            return

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"ssh_execution_log_{timestamp}.txt"

        # Get Downloads folder path
        downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
        if not os.path.exists(downloads_folder):
            downloads_folder = os.path.expanduser("~")

        # Save directly to Downloads folder
        filename = os.path.join(downloads_folder, default_filename)

        try:
            done = sum(1 for r in self.all_results if r['success'])
            failed = len(self.all_results) - done

            with open(filename, 'w', encoding='utf-8') as f:
                f.write("=" * 70 + "\n")
                f.write("       MULTIHOST SSH EXECUTION LOG\n")
                f.write("=" * 70 + "\n\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Total Hosts: {len(self.all_results)}\n")
                f.write(f"Successful: {done}\n")
                f.write(f"Failed: {failed}\n")
                f.write("=" * 70 + "\n\n")

                for result in self.all_results:
                    status = "SUCCESS" if result['success'] else f"FAILED ({result['status']})"
                    f.write(f"\n{'='*50}\n")
                    f.write(f"HOST: {result['hostname']}\n")
                    f.write(f"STATUS: {status}\n")
                    f.write(f"TIME: {result['timestamp']}\n")
                    f.write(f"{'='*50}\n")

                    if result.get('commands_output'):
                        for cmd_data in result['commands_output']:
                            f.write(f"\n$ {cmd_data['command']}\n")
                            f.write("-" * 40 + "\n")
                            f.write(f"{cmd_data['output']}\n")
                            if cmd_data.get('error'):
                                f.write(f"[STDERR]: {cmd_data['error']}\n")

                    if result.get('error'):
                        f.write(f"\nERROR: {result['error']}\n")

                f.write("\n" + "=" * 70 + "\n")
                f.write("END OF LOG\n")
                f.write("=" * 70 + "\n")

            messagebox.showinfo("Downloaded", f"Log saved to Downloads folder!\n\nFile: {default_filename}\n\nPath: {filename}")
            # Try to open the Downloads folder
            try:
                os.startfile(downloads_folder)
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {str(e)}")

    def show_download_button(self):
        """Show the download button after execution"""
        self.download_frame.pack(fill=tk.X, pady=5, before=self.broadcast_outer if hasattr(self, 'broadcast_outer') else None)
        # Update the sublabel with count
        count = len(self.all_results)
        self.download_sublabel.config(text=f"Save {count} host(s) execution log")

    def hide_download_button(self):
        """Hide the download button"""
        self.download_frame.pack_forget()

    def _build_empty_state(self):
        """Friendly placeholder shown when there are no results yet."""
        self.empty_label = tk.Frame(self.results_frame, bg=Theme.BG_DARK)
        self.empty_label.pack(pady=50)

        tk.Label(self.empty_label, text="📡", font=('Segoe UI', 42),
                 bg=Theme.BG_DARK, fg=Theme.PRIMARY).pack(pady=(0, 4))
        tk.Label(self.empty_label, text="Ready when you are",
                 font=('Segoe UI', 16, 'bold'),
                 bg=Theme.BG_DARK, fg=Theme.TEXT_WHITE).pack()
        tk.Label(self.empty_label,
                 text="Three quick steps to run a command on every host:",
                 font=('Segoe UI', 10),
                 bg=Theme.BG_DARK, fg=Theme.TEXT_MUTED).pack(pady=(6, 12))

        steps = [
            ("1", "Enter your username and password (or pick an SSH key).", Theme.CYAN),
            ("2", "Paste your hosts — one per line — into the HOSTS box.", Theme.PINK),
            ("3", "Type commands and click EXECUTE. Results appear here.", Theme.SUCCESS),
        ]
        for num, text, color in steps:
            row = tk.Frame(self.empty_label, bg=Theme.BG_DARK)
            row.pack(anchor='w', padx=20, pady=3)
            badge = tk.Frame(row, bg=color, width=24, height=24)
            badge.pack(side=tk.LEFT, padx=(0, 10))
            badge.pack_propagate(False)
            tk.Label(badge, text=num, font=('Segoe UI', 10, 'bold'),
                     bg=color, fg=Theme.TEXT_WHITE).place(relx=0.5, rely=0.5, anchor='center')
            tk.Label(row, text=text, font=('Segoe UI', 10),
                     bg=Theme.BG_DARK, fg=Theme.TEXT_LIGHT).pack(side=tk.LEFT)

        tk.Label(self.empty_label,
                 text="Tip: Press Ctrl+Enter to execute, Esc to stop.",
                 font=('Segoe UI', 9, 'italic'),
                 bg=Theme.BG_DARK, fg=Theme.TEXT_MUTED).pack(pady=(14, 0))

    def toggle_admin_mode(self):
        """Lock or unlock admin mode (controls source-file write permission)."""
        if self.admin_unlocked:
            self._set_source_readonly(True)
            self.admin_unlocked = False
            self._update_admin_btn()
            messagebox.showinfo(
                "Admin Mode",
                "Admin mode locked.\nSource file is now read-only."
            )
            return

        dialog = AdminLoginDialog(self.root, self.auth_manager)
        self.root.wait_window(dialog.win)

        if not dialog.authenticated:
            return

        # Make source file writable so the developer can edit it
        ok, msg = self._set_source_readonly(False)
        self.admin_unlocked = True
        self._update_admin_btn()

        if ok:
            messagebox.showinfo(
                "Admin Mode",
                "Admin mode unlocked.\n\nSource file is now editable.\n"
                "Click '🔓 Admin' again to lock it back."
            )
        else:
            messagebox.showwarning(
                "Admin Mode",
                f"Admin mode unlocked, but file permission change failed:\n{msg}"
            )

    def _update_admin_btn(self):
        if self.admin_unlocked:
            self.admin_btn.config(text="🔓 Unlocked", bg="#10b981",
                                  activebackground="#34d399")
        else:
            self.admin_btn.config(text="🔒 Admin", bg="#475569",
                                  activebackground="#64748b")

    def _source_path(self):
        """Resolve the source .py file path (None when running as a frozen exe)."""
        if getattr(sys, 'frozen', False):
            return None
        try:
            return os.path.abspath(__file__)
        except NameError:
            return None

    def _set_source_readonly(self, readonly):
        """Set or clear the read-only flag on the source file. Returns (ok, msg)."""
        path = self._source_path()
        if not path or not os.path.exists(path):
            return (False, "Source file not found (running as packaged exe?)")
        try:
            import stat
            if readonly:
                # Strip write bits for owner/group/other
                mode = os.stat(path).st_mode & ~(stat.S_IWUSR | stat.S_IWGRP | stat.S_IWOTH)
            else:
                mode = os.stat(path).st_mode | stat.S_IWUSR
            os.chmod(path, mode)
            return (True, "")
        except Exception as e:
            return (False, str(e))

    def animate_footer_glow(self):
        """Animate the footer glow bar"""
        colors = ["#667eea", "#764ba2", "#ec4899", "#f97316", "#10b981", "#06b6d4"]
        if not hasattr(self, '_footer_color_idx'):
            self._footer_color_idx = 0
        self._footer_color_idx = (self._footer_color_idx + 1) % len(colors)
        try:
            self.footer_glow.config(bg=colors[self._footer_color_idx])
        except Exception:
            pass
        self.root.after(600, self.animate_footer_glow)

    def create_section_header(self, parent, title, subtitle=""):
        frame = tk.Frame(parent, bg=Theme.BG_MAIN)
        frame.pack(fill=tk.X, pady=10)

        tk.Label(frame, text=title, font=('Segoe UI', 11, 'bold'),
                bg=Theme.BG_MAIN, fg=Theme.TEXT_WHITE).pack(anchor=tk.W)

        if subtitle:
            tk.Label(frame, text=subtitle, font=('Segoe UI', 9),
                    bg=Theme.BG_MAIN, fg=Theme.TEXT_MUTED).pack(anchor=tk.W)

    def create_card(self, parent, accent_color=None):
        """Create a card with optional accent color border"""
        # Outer frame for border effect
        border_color = accent_color if accent_color else Theme.BORDER
        outer = tk.Frame(parent, bg=border_color)
        outer.pack(fill=tk.X, pady=4)

        # Inner card
        card = tk.Frame(outer, bg=Theme.BG_CARD)
        card.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)

        # Top accent bar if color provided
        if accent_color:
            accent_bar = tk.Frame(card, bg=accent_color, height=3)
            accent_bar.pack(fill=tk.X)

        return card

    def create_input_field(self, parent, label):
        tk.Label(parent, text=label, font=('Segoe UI', 9),
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor=tk.W, padx=12, pady=2)

    def toggle_password(self):
        self.pass_entry.config(show="" if self.show_pass_var.get() else "*")

    def toggle_auth_method(self):
        if self.auth_method_var.get() == "password":
            self.key_frame.pack_forget()
            self.password_frame.pack(fill=tk.X)
        else:
            self.password_frame.pack_forget()
            self.key_frame.pack(fill=tk.X)

    def browse_key_file(self):
        file_path = filedialog.askopenfilename(
            title="Select SSH Key File",
            filetypes=[("All files", "*.*"), ("PEM files", "*.pem"), ("Key files", "*.key")]
        )
        if file_path:
            self.key_file_entry.delete(0, tk.END)
            self.key_file_entry.insert(0, file_path)

    def update_progress(self, progress):
        width = self.progress_bar.master.winfo_width()
        bar_width = int(width * progress / 100)
        color = Theme.SUCCESS if progress == 100 else Theme.PRIMARY
        self.progress_bar.config(width=bar_width, bg=color)

    def _get_auth_params(self):
        """Return (password, key_file, key_passphrase) tuple based on selected auth method."""
        if self.auth_method_var.get() == "password":
            return (self.pass_entry.get(), None, None)
        return (
            None,
            self.key_file_entry.get().strip(),
            self.key_passphrase_entry.get() or None,
        )

    def get_hosts(self):
        return [l.strip() for l in self.hosts_text.get("1.0", tk.END).split('\n')
                if l.strip() and not l.strip().startswith('#')]

    def get_commands(self):
        return [l.strip() for l in self.commands_text.get("1.0", tk.END).split('\n')
                if l.strip() and not l.strip().startswith('#')]

    def validate_inputs(self):
        if not self.get_hosts():
            messagebox.showerror("Error", "Enter at least one host")
            return False
        if not self.user_entry.get().strip():
            messagebox.showerror("Error", "Enter username")
            return False

        if self.auth_method_var.get() == "password":
            if not self.pass_entry.get():
                messagebox.showerror("Error", "Enter password")
                return False
        else:
            key_file = self.key_file_entry.get().strip()
            if not key_file or not os.path.exists(key_file):
                messagebox.showerror("Error", "Select a valid SSH key file")
                return False

        if not self.get_commands():
            messagebox.showerror("Error", "Enter at least one command")
            return False
        try:
            port = int(self.port_entry.get())
            if not 1 <= port <= 65535:
                raise ValueError
        except Exception:
            messagebox.showerror("Error", "Invalid port")
            return False
        try:
            max_concurrent = int(self.max_concurrent_entry.get())
            if not 1 <= max_concurrent <= 100:
                raise ValueError
        except Exception:
            messagebox.showerror("Error", "Invalid max concurrent (1-100)")
            return False
        timeout_raw = self.timeout_entry.get().strip()
        if timeout_raw:
            try:
                timeout = int(timeout_raw)
                if not 1 <= timeout <= 3600:
                    raise ValueError
            except Exception:
                messagebox.showerror("Error", "Invalid timeout (1-3600 seconds)")
                return False
        return True

    def execute_commands(self):
        if self.is_executing:
            return
        if not self.validate_inputs():
            return

        self.clear_results()

        hosts = self.get_hosts()
        commands = self.get_commands()
        username = self.user_entry.get().strip()
        port = int(self.port_entry.get())
        timeout = int(self.timeout_entry.get() or "30")
        max_concurrent = int(self.max_concurrent_entry.get() or "10")

        password, key_file, key_passphrase = self._get_auth_params()

        self.is_executing = True
        self.cancel_requested = False
        self.executors = []
        self.execute_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.empty_label.pack_forget()

        for host in hosts:
            card = HostResultCard(self.results_frame, host)
            card.pack(fill=tk.X, padx=8, pady=6)
            self.host_cards[host] = card

        self.total_label.config(text=f" TOTAL: {len(hosts)} ")
        self.status_label.config(text=f"Executing on {len(hosts)} host(s)...", fg=Theme.INFO_LIGHT)

        thread = threading.Thread(
            target=self.run_execution,
            args=(hosts, commands, username, password, port, timeout, key_file, key_passphrase, max_concurrent),
            daemon=True
        )
        thread.start()

    def run_execution(self, hosts, commands, username, password, port, timeout,
                     key_file, key_passphrase, max_concurrent):
        total = len(hosts)
        completed = 0

        with ThreadPoolExecutor(max_workers=min(max_concurrent, total)) as executor:
            futures = {}
            for host in hosts:
                ssh_executor = SSHExecutor(
                    host, username, password, port, timeout,
                    key_file, key_passphrase
                )
                self.executors.append(ssh_executor)
                future = executor.submit(ssh_executor.execute_commands, commands)
                futures[future] = host

            for future in as_completed(futures):
                if self.cancel_requested:
                    # Cancel remaining executors
                    for exc in self.executors:
                        exc.cancel()

                result = future.result()
                completed += 1
                progress = (completed / total) * 100
                self.result_queue.put(('result', result, progress, completed, total))

        self.result_queue.put(('done', None, 100, total, total))

    def stop_execution(self):
        """Stop the current execution"""
        if self.is_executing:
            self.cancel_requested = True
            for executor in self.executors:
                executor.cancel()
            self.status_label.config(text="Stopping execution...", fg=Theme.WARNING)

    def rerun_failed(self):
        """Re-run commands on failed hosts"""
        if self.is_executing:
            messagebox.showwarning("Warning", "Execution in progress")
            return

        failed_hosts = [r['hostname'] for r in self.all_results if not r['success']]

        if not failed_hosts:
            messagebox.showinfo("Info", "No failed hosts to retry")
            return

        # Update hosts text with failed hosts only
        self.hosts_text.delete('1.0', tk.END)
        self.hosts_text.insert(tk.END, '\n'.join(failed_hosts))

        if messagebox.askyesno("Retry Failed Hosts",
                               f"Retry execution on {len(failed_hosts)} failed host(s)?"):
            self.execute_commands()

    def broadcast_command(self):
        """Send a command to all hosts simultaneously"""
        command = self.broadcast_entry.get().strip()

        if not command:
            messagebox.showerror("Error", "Enter a command to broadcast")
            return

        if self.is_executing:
            messagebox.showwarning("Warning", "Execution already in progress")
            return

        if not self.all_results:
            messagebox.showinfo("Info", "No hosts to send command to.\nExecute commands first.")
            return

        # Get target hosts based on checkbox
        if self.broadcast_success_only_var.get():
            target_hosts = [r['hostname'] for r in self.all_results if r['success']]
        else:
            target_hosts = [r['hostname'] for r in self.all_results]

        if not target_hosts:
            messagebox.showinfo("Info", "No target hosts available")
            return

        # Confirm broadcast
        if not messagebox.askyesno("Broadcast Command",
                                   f"Send command to {len(target_hosts)} host(s)?\n\n"
                                   f"Command: {command}"):
            return

        # Get credentials
        username = self.user_entry.get().strip()
        port = int(self.port_entry.get())
        timeout = int(self.timeout_entry.get() or "30")
        max_concurrent = int(self.max_concurrent_entry.get() or "10")

        password, key_file, key_passphrase = self._get_auth_params()

        # Update UI
        self.is_executing = True
        self.execute_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.broadcast_btn.config(state='disabled')
        self.status_label.config(text=f"Broadcasting to {len(target_hosts)} host(s)...", fg=Theme.CYAN)

        # Clear the broadcast entry
        self.broadcast_entry.delete(0, tk.END)

        # Execute in background
        thread = threading.Thread(
            target=self.run_broadcast,
            args=(target_hosts, [command], username, password, port, timeout,
                  key_file, key_passphrase, max_concurrent),
            daemon=True
        )
        thread.start()

    def run_broadcast(self, hosts, commands, username, password, port, timeout,
                      key_file, key_passphrase, max_concurrent):
        """Execute broadcast command on all target hosts"""
        total = len(hosts)
        completed = 0
        broadcast_results = []

        with ThreadPoolExecutor(max_workers=min(max_concurrent, total)) as executor:
            futures = {}
            for host in hosts:
                ssh_executor = SSHExecutor(
                    host, username, password, port, timeout,
                    key_file, key_passphrase
                )
                self.executors.append(ssh_executor)
                future = executor.submit(ssh_executor.execute_commands, commands)
                futures[future] = host

            for future in as_completed(futures):
                if self.cancel_requested:
                    for exc in self.executors:
                        exc.cancel()

                result = future.result()
                completed += 1
                broadcast_results.append(result)

                # Update the existing card with new output appended
                self.result_queue.put(('broadcast_result', result, completed, total))

        self.result_queue.put(('broadcast_done', broadcast_results, total, total))

    def check_queue(self):
        try:
            while True:
                msg = self.result_queue.get_nowait()

                if msg[0] == 'result':
                    _, result, progress, completed, total = msg
                    self.all_results.append(result)

                    if result['hostname'] in self.host_cards:
                        self.host_cards[result['hostname']].update_result(result)

                    self.update_progress(progress)

                    done = sum(1 for r in self.all_results if r['success'])
                    failed = len(self.all_results) - done

                    self.done_label.config(text=f" SUCCESS: {done} ")
                    self.failed_label.config(text=f" FAILED: {failed} ")
                    self.status_label.config(text=f"Completed {completed}/{total}")

                elif msg[0] == 'broadcast_result':
                    _, result, completed, total = msg

                    # Append broadcast output to existing card
                    if result['hostname'] in self.host_cards:
                        card = self.host_cards[result['hostname']]
                        card.output_text.config(state=tk.NORMAL)

                        # Add separator and broadcast output
                        card.output_text.insert(tk.END, "\n\n" + "=" * 45 + "\n", 'separator')
                        card.output_text.insert(tk.END, ">>> BROADCAST COMMAND <<<\n", 'command')
                        card.output_text.insert(tk.END, f"Time: {result['timestamp']}\n", 'separator')
                        card.output_text.insert(tk.END, "-" * 45 + "\n", 'separator')

                        if result.get('commands_output'):
                            for cmd_data in result['commands_output']:
                                card.output_text.insert(tk.END, f"$ {cmd_data['command']}\n", 'command')
                                output = cmd_data.get('output', '')
                                if output:
                                    card.output_text.insert(tk.END, f"{output}\n", 'output')
                                if cmd_data.get('error'):
                                    card.output_text.insert(tk.END, f"[STDERR]: {cmd_data['error']}\n", 'error')
                        elif result.get('error'):
                            card.output_text.insert(tk.END, f"ERROR: {result['error']}\n", 'error')

                        card.output_text.see(tk.END)
                        card.output_text.config(state=tk.DISABLED)

                        # Expand the card to show new output
                        if not card.expanded:
                            card.toggle_output()

                    self.status_label.config(text=f"Broadcast: {completed}/{total} hosts", fg=Theme.CYAN)

                elif msg[0] == 'broadcast_done':
                    _, results, completed, total = msg
                    self.is_executing = False
                    self.execute_btn.config(state='normal')
                    self.stop_btn.config(state='disabled')
                    self.broadcast_btn.config(state='normal')
                    self.executors = []

                    success = sum(1 for r in results if r['success'])
                    failed = len(results) - success

                    self.status_label.config(
                        text=f"Broadcast complete: {success} Success | {failed} Failed",
                        fg=Theme.SUCCESS_LIGHT if failed == 0 else Theme.WARNING
                    )

                elif msg[0] == 'done':
                    self.is_executing = False
                    self.execute_btn.config(state='normal')
                    self.stop_btn.config(state='disabled')
                    self.broadcast_btn.config(state='normal')
                    self.executors = []

                    done = sum(1 for r in self.all_results if r['success'])
                    failed = len(self.all_results) - done
                    cancelled = sum(1 for r in self.all_results if r['status'] == 'CANCELLED')

                    if cancelled > 0:
                        self.status_label.config(text=f"Stopped: {done} Success | {failed} Failed | {cancelled} Cancelled", fg=Theme.WARNING)
                    elif failed == 0:
                        self.status_label.config(text=f"ALL {done} HOST(S) COMPLETED SUCCESSFULLY!", fg=Theme.SUCCESS_LIGHT)
                    else:
                        self.status_label.config(text=f"Completed: {done} Success | {failed} Failed", fg=Theme.WARNING)

                    # Show download button after execution completes
                    if self.all_results:
                        self.show_download_button()

        except queue.Empty:
            pass

        self.root.after(100, self.check_queue)

    def clear_results(self):
        self.all_results = []
        self.host_cards = {}

        for widget in self.results_frame.winfo_children():
            widget.destroy()

        self._build_empty_state()

        self.update_progress(0)
        self.total_label.config(text=" TOTAL: 0 ")
        self.done_label.config(text=" SUCCESS: 0 ")
        self.failed_label.config(text=" FAILED: 0 ")
        self.status_label.config(text="Ready to execute", fg=Theme.CYAN)

        # Hide download button
        self.hide_download_button()

    def copy_results_to_clipboard(self):
        """Copy all results to clipboard"""
        if not self.all_results:
            messagebox.showinfo("Info", "No results to copy")
            return

        output = []
        output.append("=" * 70)
        output.append("         MULTIHOST SSH EXECUTION RESULTS")
        output.append("=" * 70)
        output.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        done = sum(1 for r in self.all_results if r['success'])
        failed = len(self.all_results) - done
        output.append(f"Total: {len(self.all_results)} | Success: {done} | Failed: {failed}")
        output.append("=" * 70)

        for result in self.all_results:
            status = "SUCCESS" if result['success'] else "FAILED"
            output.append(f"\n[{status}] {result['hostname']}")
            output.append(f"Time: {result['timestamp']}")
            output.append("-" * 50)

            if result.get('commands_output'):
                for cmd_data in result['commands_output']:
                    output.append(f"\n$ {cmd_data['command']}")
                    output.append(cmd_data['output'])
                    if cmd_data['error']:
                        output.append(f"[STDERR]: {cmd_data['error']}")

            if result['error']:
                output.append(f"\nERROR: {result['error']}")

            output.append("-" * 50)

        text = '\n'.join(output)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        messagebox.showinfo("Copied", "Results copied to clipboard")

    def import_hosts(self):
        """Import hosts from TXT or CSV file"""
        file_path = filedialog.askopenfilename(
            title="Import Hosts",
            filetypes=[("Text files", "*.txt"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )

        if not file_path:
            return

        hosts = []
        try:
            if file_path.lower().endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if row:
                            # Take first column as hostname
                            host = row[0].strip()
                            if host and not host.startswith('#'):
                                hosts.append(host)
            else:
                with open(file_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        host = line.strip()
                        if host and not host.startswith('#'):
                            hosts.append(host)

            if hosts:
                self.hosts_text.delete('1.0', tk.END)
                self.hosts_text.insert(tk.END, '\n'.join(hosts))
                messagebox.showinfo("Imported", f"Imported {len(hosts)} host(s)")
            else:
                messagebox.showwarning("Warning", "No valid hosts found in file")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to import: {str(e)}")

    def save_session(self):
        """Save current session configuration"""
        config = {
            'username': self.user_entry.get(),
            'port': self.port_entry.get(),
            'timeout': self.timeout_entry.get(),
            'max_concurrent': self.max_concurrent_entry.get(),
            'auth_method': self.auth_method_var.get(),
            'key_file': self.key_file_entry.get(),
            'hosts': self.get_hosts(),
            'commands': self.get_commands()
        }

        if self.session_manager.save_session(config):
            messagebox.showinfo("Saved", "Session configuration saved")
        else:
            messagebox.showerror("Error", "Failed to save session")

    def load_session(self):
        """Load saved session configuration"""
        config = self.session_manager.load_session()

        if config:
            if config.get('username'):
                self.user_entry.delete(0, tk.END)
                self.user_entry.insert(0, config['username'])
            if config.get('port'):
                self.port_entry.delete(0, tk.END)
                self.port_entry.insert(0, config['port'])
            if config.get('timeout'):
                self.timeout_entry.delete(0, tk.END)
                self.timeout_entry.insert(0, config['timeout'])
            if config.get('max_concurrent'):
                self.max_concurrent_entry.delete(0, tk.END)
                self.max_concurrent_entry.insert(0, config['max_concurrent'])
            if config.get('auth_method'):
                self.auth_method_var.set(config['auth_method'])
                self.toggle_auth_method()
            if config.get('key_file'):
                self.key_file_entry.delete(0, tk.END)
                self.key_file_entry.insert(0, config['key_file'])
            if config.get('hosts'):
                self.hosts_text.delete('1.0', tk.END)
                self.hosts_text.insert(tk.END, '\n'.join(config['hosts']))
            if config.get('commands'):
                self.commands_text.delete('1.0', tk.END)
                self.commands_text.insert(tk.END, '\n'.join(config['commands']))

    # ============== HOST GROUP METHODS ==============
    def update_host_group_list(self):
        groups = self.host_group_manager.get_group_names()
        self.host_group_combo['values'] = groups
        self.host_group_combo.set('')

    def load_host_group(self):
        selected = self.host_group_var.get()
        if not selected:
            messagebox.showinfo("Info", "Select a host group to load")
            return

        hosts = self.host_group_manager.get_group(selected)
        if hosts:
            self.hosts_text.delete('1.0', tk.END)
            self.hosts_text.insert(tk.END, '\n'.join(hosts))
            messagebox.showinfo("Loaded", f"Loaded host group: {selected}")
        else:
            messagebox.showwarning("Warning", f"Host group '{selected}' is empty")

    def save_host_group(self):
        hosts = self.get_hosts()
        if not hosts:
            messagebox.showerror("Error", "Enter hosts before saving")
            return

        name = simpledialog.askstring("Save Host Group",
                                      "Enter a name for this host group:",
                                      parent=self.root)
        if not name:
            return

        name = name.strip()
        if not name:
            messagebox.showerror("Error", "Group name cannot be empty")
            return

        if name in self.host_group_manager.get_group_names():
            if not messagebox.askyesno("Confirm",
                                       f"Host group '{name}' already exists.\nOverwrite?"):
                return

        if self.host_group_manager.add_group(name, hosts):
            self.update_host_group_list()
            self.host_group_var.set(name)
            messagebox.showinfo("Saved", f"Host group '{name}' saved!")
        else:
            messagebox.showerror("Error", "Failed to save host group")

    def delete_host_group(self):
        selected = self.host_group_var.get()
        if not selected:
            messagebox.showinfo("Info", "Select a host group to delete")
            return

        if messagebox.askyesno("Confirm Delete",
                               f"Delete host group '{selected}'?"):
            if self.host_group_manager.delete_group(selected):
                self.update_host_group_list()
                self.host_group_var.set('')
                messagebox.showinfo("Deleted", f"Host group '{selected}' deleted")
            else:
                messagebox.showerror("Error", "Failed to delete host group")

    def export_results(self):
        if not self.all_results:
            messagebox.showinfo("Info", "No results to export")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt")],
            initialfilename=f"ssh_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )

        if filename:
            try:
                done = sum(1 for r in self.all_results if r['success'])
                failed = len(self.all_results) - done

                with open(filename, 'w', encoding='utf-8') as f:
                    f.write("=" * 70 + "\n")
                    f.write("         MULTIHOST SSH EXECUTION RESULTS\n")
                    f.write("=" * 70 + "\n\n")
                    f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"Total: {len(self.all_results)} | Success: {done} | Failed: {failed}\n")
                    f.write("=" * 70 + "\n")

                    for result in self.all_results:
                        status = "SUCCESS" if result['success'] else "FAILED"
                        f.write(f"\n[{status}] {result['hostname']}\n")
                        f.write(f"Time: {result['timestamp']}\n")
                        f.write("-" * 50 + "\n")

                        if result.get('commands_output'):
                            for cmd_data in result['commands_output']:
                                f.write(f"\n$ {cmd_data['command']}\n")
                                f.write(f"{cmd_data['output']}\n")
                                if cmd_data['error']:
                                    f.write(f"[STDERR]: {cmd_data['error']}\n")

                        if result['error']:
                            f.write(f"\nERROR: {result['error']}\n")

                        f.write("-" * 50 + "\n")

                messagebox.showinfo("Success", f"Saved to:\n{filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Could not save: {str(e)}")

    def export_to_excel(self):
        """Export results to Excel file directly to Downloads folder"""
        if not self.all_results:
            messagebox.showinfo("Info", "No results to export")
            return

        if not EXCEL_AVAILABLE:
            messagebox.showerror("Error",
                "Excel export requires 'openpyxl' library.\n\n"
                "Install it using:\npip install openpyxl")
            return

        # Save directly to Downloads folder
        downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
        if not os.path.exists(downloads_folder):
            downloads_folder = os.path.expanduser("~")

        default_filename = f"ssh_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filename = os.path.join(downloads_folder, default_filename)

        try:
            wb = Workbook()

            # ===== SUMMARY SHEET =====
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
            success_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            success_font = Font(bold=True, color="166534")
            error_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            error_font = Font(bold=True, color="991B1B")
            border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Title
            ws_summary['A1'] = "SSH Multi-Host Execution Results"
            ws_summary['A1'].font = Font(bold=True, size=16, color="1E293B")
            ws_summary.merge_cells('A1:E1')

            # Stats
            done_count = sum(1 for r in self.all_results if r['success'])
            failed_count = len(self.all_results) - done_count

            ws_summary['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws_summary['A4'] = f"Total Hosts: {len(self.all_results)}"
            ws_summary['A5'] = f"Successful: {done_count}"
            ws_summary['A5'].font = success_font
            ws_summary['A6'] = f"Failed: {failed_count}"
            ws_summary['A6'].font = error_font

            # Summary table headers
            headers = ["#", "Hostname", "Status", "Timestamp", "Error"]
            for col, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=8, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_align

            # Summary data
            for idx, result in enumerate(self.all_results, 1):
                row = 8 + idx
                ws_summary.cell(row=row, column=1, value=idx).border = border
                ws_summary.cell(row=row, column=2, value=result['hostname']).border = border

                status_cell = ws_summary.cell(row=row, column=3, value="SUCCESS" if result['success'] else result['status'])
                status_cell.border = border
                status_cell.alignment = center_align
                if result['success']:
                    status_cell.fill = success_fill
                    status_cell.font = success_font
                else:
                    status_cell.fill = error_fill
                    status_cell.font = error_font

                ws_summary.cell(row=row, column=4, value=result['timestamp']).border = border
                ws_summary.cell(row=row, column=5, value=result.get('error', '')).border = border

            # Adjust column widths for summary
            ws_summary.column_dimensions['A'].width = 5
            ws_summary.column_dimensions['B'].width = 25
            ws_summary.column_dimensions['C'].width = 15
            ws_summary.column_dimensions['D'].width = 20
            ws_summary.column_dimensions['E'].width = 40

            # ===== DETAILED RESULTS SHEET =====
            ws_details = wb.create_sheet("Detailed Results")

            # Headers
            detail_headers = ["Hostname", "Status", "Command", "Output", "Error", "Timestamp"]
            for col, header in enumerate(detail_headers, 1):
                cell = ws_details.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_align

            # Detail data
            detail_row = 2
            for result in self.all_results:
                hostname = result['hostname']
                status = "SUCCESS" if result['success'] else result['status']
                timestamp = result['timestamp']

                if result.get('commands_output'):
                    for cmd_data in result['commands_output']:
                        ws_details.cell(row=detail_row, column=1, value=hostname).border = border

                        status_cell = ws_details.cell(row=detail_row, column=2, value=status)
                        status_cell.border = border
                        if result['success']:
                            status_cell.fill = success_fill
                            status_cell.font = success_font
                        else:
                            status_cell.fill = error_fill
                            status_cell.font = error_font

                        ws_details.cell(row=detail_row, column=3, value=cmd_data['command']).border = border

                        output_cell = ws_details.cell(row=detail_row, column=4, value=cmd_data['output'])
                        output_cell.border = border
                        output_cell.alignment = wrap_align

                        ws_details.cell(row=detail_row, column=5, value=cmd_data.get('error', '')).border = border
                        ws_details.cell(row=detail_row, column=6, value=timestamp).border = border

                        detail_row += 1
                else:
                    # No commands output - connection error
                    ws_details.cell(row=detail_row, column=1, value=hostname).border = border

                    status_cell = ws_details.cell(row=detail_row, column=2, value=status)
                    status_cell.border = border
                    status_cell.fill = error_fill
                    status_cell.font = error_font

                    ws_details.cell(row=detail_row, column=3, value="N/A").border = border
                    ws_details.cell(row=detail_row, column=4, value="N/A").border = border
                    ws_details.cell(row=detail_row, column=5, value=result.get('error', '')).border = border
                    ws_details.cell(row=detail_row, column=6, value=timestamp).border = border

                    detail_row += 1

            # Adjust column widths for details
            ws_details.column_dimensions['A'].width = 20
            ws_details.column_dimensions['B'].width = 12
            ws_details.column_dimensions['C'].width = 30
            ws_details.column_dimensions['D'].width = 50
            ws_details.column_dimensions['E'].width = 30
            ws_details.column_dimensions['F'].width = 20

            # Save workbook
            wb.save(filename)
            messagebox.showinfo("Success", f"Excel file saved to Downloads folder!\n\nFile: {default_filename}\n\nPath: {filename}")
            # Try to open the Downloads folder
            try:
                os.startfile(downloads_folder)
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror("Error", f"Could not save Excel file:\n{str(e)}")

    # ============== PRESET METHODS ==============
    def update_preset_list(self):
        presets = self.preset_manager.get_preset_names()
        self.preset_combo['values'] = presets
        self.preset_combo.set('')

    def on_preset_selected(self, event=None):
        selected = self.preset_var.get()
        if selected:
            self.load_preset_commands(selected)

    def load_preset(self):
        selected = self.preset_var.get()
        if not selected:
            messagebox.showinfo("Info", "Select a preset to load")
            return
        self.load_preset_commands(selected)

    def load_preset_commands(self, preset_name):
        commands = self.preset_manager.get_preset(preset_name)
        if commands:
            self.commands_text.delete('1.0', tk.END)
            self.commands_text.insert(tk.END, '\n'.join(commands))
            messagebox.showinfo("Loaded", f"Loaded preset: {preset_name}")
        else:
            messagebox.showwarning("Warning", f"Preset '{preset_name}' is empty")

    def save_preset(self):
        commands = self.get_commands()
        if not commands:
            messagebox.showerror("Error", "Enter commands before saving")
            return

        name = simpledialog.askstring("Save Preset",
                                      "Enter a name for this preset:",
                                      parent=self.root)
        if not name:
            return

        name = name.strip()
        if not name:
            messagebox.showerror("Error", "Preset name cannot be empty")
            return

        if name in self.preset_manager.get_preset_names():
            if not messagebox.askyesno("Confirm",
                                       f"Preset '{name}' already exists.\nOverwrite?"):
                return

        if self.preset_manager.add_preset(name, commands):
            self.update_preset_list()
            self.preset_var.set(name)
            messagebox.showinfo("Saved", f"Preset '{name}' saved successfully!")
        else:
            messagebox.showerror("Error", "Failed to save preset")

    def delete_preset(self):
        selected = self.preset_var.get()
        if not selected:
            messagebox.showinfo("Info", "Select a preset to delete")
            return

        if messagebox.askyesno("Confirm Delete",
                               f"Delete preset '{selected}'?"):
            if self.preset_manager.delete_preset(selected):
                self.update_preset_list()
                self.preset_var.set('')
                messagebox.showinfo("Deleted", f"Preset '{selected}' deleted")
            else:
                messagebox.showerror("Error", "Failed to delete preset")


def main():
    root = tk.Tk()
    app = MultiHostExecutorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
